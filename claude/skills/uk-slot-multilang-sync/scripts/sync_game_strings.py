#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""多國語言文字同步工具(線上部門 uk 老虎機框架)。

讀 xlsx 多國語言工作表 + 對應表(Key ↔ 列號),
把指定列的各國文字 upsert 進 <text-dir>/<code>/gameStrings.xml。

慣例假設(uk 線上老虎機框架,跨專案穩定):
  - gameStrings.xml 是 plist,每個 Key 是一個標籤 <KEY>文字</KEY>。
  - 每個語系一個目錄,目錄名 = 代碼(tw/cn/en/...);語系數量依專案(常見 26)。
  - xlsx「多國語言」工作表:列=Key、欄=語系,語系欄標題括號內就是代碼,如「繁體中文(tw)zh-TW」→ tw。

關鍵紀律(寫進 SKILL.md,這裡再強制一次):
  - 「列號」= xlsx 的「實際 Excel 列號」,不是內部編號欄。
  - 預設 dry-run 只印不寫;--write 才實際寫入。
  - 保留各 XML 原本的換行(CRLF/LF)與 UTF-8(無 BOM),文字做 XML 跳脫(& < >)。

路徑全部相對「當前工作目錄」(在目標專案根目錄執行),不依賴本檔位置,
因此放在 skill 目錄也能正確操作目標專案。

用法(在專案根目錄):
    python <skill>/scripts/sync_game_strings.py --xlsx 規格.xlsx           # dry-run
    python <skill>/scripts/sync_game_strings.py --xlsx 規格.xlsx --write   # 實際寫入
    可選:--mapping 路徑(預設 ./string_keys.tsv)
          --text-dir 路徑(預設 ./assets/game/Text)
          --sheet 名稱(預設自動偵測含「多國語言/多語」的工作表)
"""
import argparse
import re
import sys
from pathlib import Path

import openpyxl

FIRST_LANG_COL_IDX0 = 2  # 0-based:前兩欄(分類/編號)之後才是語言欄
SHEET_HINTS = ("多國語言", "多国语言", "多語", "多语", "multilang", "multi-lang", "i18n")


def xml_escape(s: str) -> str:
    """元素內容跳脫;& 必須先處理。"""
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def pick_sheet(wb, explicit):
    """選工作表:有 --sheet 用它;否則自動偵測含關鍵字者(需唯一)。"""
    if explicit:
        if explicit in wb.sheetnames:
            return explicit
        print(f"❌ 找不到工作表 {explicit!r};可用:{wb.sheetnames}")
        sys.exit(1)
    cands = [s for s in wb.sheetnames if any(h in s.lower() for h in (x.lower() for x in SHEET_HINTS))]
    if len(cands) == 1:
        print(f"🔎 自動選定工作表:{cands[0]!r}")
        return cands[0]
    if not cands:
        print(f"❌ 找不到含『多國語言』的工作表,請用 --sheet 指定;可用:{wb.sheetnames}")
        sys.exit(1)
    print(f"❌ 多個工作表符合 {cands},請用 --sheet 明確指定")
    sys.exit(1)


def parse_header(header_row):
    """第 1 列 → 代碼(括號內) → 0-based 欄索引。"""
    code_to_idx = {}
    for idx, cell in enumerate(header_row):
        if idx < FIRST_LANG_COL_IDX0 or cell is None:
            continue
        m = re.search(r"\(([^)]+)\)", str(cell))
        if m:
            code_to_idx[m.group(1).strip()] = idx
    return code_to_idx


def load_mapping(path: Path):
    """每行 Key<Tab>列號<Tab>備註(可選);# 開頭與空行忽略。"""
    entries = []
    for lineno, raw in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in raw.split("\t") if p.strip() != ""]
        if len(parts) < 2:
            print(f"  ⚠️ 對應表第 {lineno} 行格式不符(需 Key<Tab>列號),略過:{raw!r}")
            continue
        key, rownum = parts[0], parts[1]
        note = parts[2] if len(parts) >= 3 else ""
        if not re.fullmatch(r"[A-Za-z0-9_]+", key):
            print(f"  ⚠️ 對應表第 {lineno} 行 Key 含非法字元,略過:{key!r}")
            continue
        if not rownum.isdigit():
            print(f"  ⚠️ 對應表第 {lineno} 行列號非數字,略過:{rownum!r}")
            continue
        entries.append((key, int(rownum), note))
    return entries


def cell_to_text(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def upsert(text: str, key: str, value: str, nl: str):
    """回傳 (新文字, 'updated'|'added')。"""
    esc = xml_escape(value).replace("\r\n", "\n").replace("\n", nl)
    pat = re.compile(r"(<%s>)(.*?)(</%s>)" % (re.escape(key), re.escape(key)), re.DOTALL)
    if pat.search(text):
        new = pat.sub(lambda m: m.group(1) + esc + m.group(3), text, count=1)
        return new, "updated"
    entry = "\t\t<%s>%s</%s>" % (key, esc, key)
    m = re.compile(r"(\r?\n)([ \t]*</dict>)").search(text)
    if not m:
        raise ValueError("找不到 </dict>,無法插入新 Key")
    insert_at = m.start()
    return text[:insert_at] + nl + entry + text[insert_at:], "added"


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # Windows cp950 印不出 emoji
    except Exception:
        pass

    ap = argparse.ArgumentParser(description="同步多國語言文字到各語系 gameStrings.xml")
    ap.add_argument("--xlsx", type=Path, required=True, help="規格 xlsx 路徑")
    ap.add_argument("--mapping", type=Path, default=Path("string_keys.tsv"), help="Key↔列號對應表(預設 ./string_keys.tsv)")
    ap.add_argument("--text-dir", type=Path, default=Path("assets/game/Text"), help="語系目錄根(預設 ./assets/game/Text)")
    ap.add_argument("--xml-name", default="gameStrings.xml", help="各語系內的 XML 檔名")
    ap.add_argument("--sheet", default=None, help="工作表名(預設自動偵測含『多國語言』者)")
    ap.add_argument("--write", action="store_true", help="實際寫入(預設 dry-run)")
    args = ap.parse_args()

    for label, p in (("xlsx", args.xlsx), ("mapping", args.mapping)):
        if not p.exists():
            print(f"❌ 找不到 {label}:{p}(請確認在專案根目錄執行,或用參數指定路徑)")
            return 1
    if not args.text_dir.is_dir():
        print(f"❌ 找不到語系目錄:{args.text_dir}(用 --text-dir 指定)")
        return 1

    mapping = load_mapping(args.mapping)
    if not mapping:
        print("⚠️ 對應表沒有任何有效條目,無事可做。")
        return 0

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    sheet = pick_sheet(wb, args.sheet)
    ws = wb[sheet]

    needed = {row for _, row, _ in mapping}
    header_row = None
    rows_data = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            header_row = row
        if i in needed:
            rows_data[i] = row
        if header_row is not None and needed.issubset(rows_data.keys()) and i >= max(needed):
            break

    code_to_idx = parse_header(header_row)
    missing_dirs = [c for c in code_to_idx if not (args.text_dir / c / args.xml_name).exists()]
    if missing_dirs:
        print(f"❌ 下列語系代碼找不到對應 {args.xml_name},請檢查目錄/欄標題:{missing_dirs}")
        return 1

    print(f"📄 xlsx    : {args.xlsx}  [工作表 {sheet!r}]")
    print(f"📄 mapping : {args.mapping}  ({len(mapping)} 筆 Key)")
    print(f"🌐 語系    : {len(code_to_idx)} 個 → {sorted(code_to_idx)}")
    print(f"✍️  模式    : {'WRITE 實際寫入' if args.write else 'DRY-RUN 只預覽不寫'}")
    print("=" * 70)

    file_text, file_nl = {}, {}
    stats = {"added": 0, "updated": 0, "missing": 0}

    for key, rownum, note in mapping:
        if rownum not in rows_data:
            print(f"❌ [{key}] 列號 {rownum} 超出資料範圍,略過")
            continue
        row = rows_data[rownum]
        added, updated, missing = [], [], []
        for code, idx in code_to_idx.items():
            value = cell_to_text(row[idx]) if idx < len(row) else None
            if value is None or value == "":
                missing.append(code)
                continue
            xml_path = args.text_dir / code / args.xml_name
            if code not in file_text:
                raw = xml_path.read_bytes().decode("utf-8")
                file_text[code] = raw
                file_nl[code] = "\r\n" if "\r\n" in raw else "\n"
            new_text, action = upsert(file_text[code], key, value, file_nl[code])
            file_text[code] = new_text
            (added if action == "added" else updated).append(code)
        stats["added"] += len(added)
        stats["updated"] += len(updated)
        stats["missing"] += len(missing)
        tag = f" ({note})" if note else ""
        print(f"🔑 {key}  ← 列 {rownum}{tag}")
        print(f"   ＋新增 {len(added):2d} | ⟳更新 {len(updated):2d} | ∅缺字 {len(missing):2d}")
        if missing:
            print(f"   ∅ 缺翻譯語系:{missing}")

    print("=" * 70)
    print(f"合計:新增 {stats['added']} | 更新 {stats['updated']} | 缺字 {stats['missing']}")

    if not args.write:
        print("\nℹ️ 這是 DRY-RUN,未寫入任何檔案。確認無誤後加 --write 實際執行。")
        return 0

    for code, text in file_text.items():
        (args.text_dir / code / args.xml_name).write_bytes(text.encode("utf-8"))
    print(f"\n✅ 已寫入 {len(file_text)} 個 {args.xml_name}。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
