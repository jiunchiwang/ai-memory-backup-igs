"""
Excel → AI Document 轉換 pipeline

用途：把含大量示意圖的遊戲規格書(.xlsx) 轉成 AI 可理解的知識結構
輸出：markdown/ + images/ + metadata/ + source/

核心策略：
1. MarkItDown 抽文字/表格
2. 解 xlsx zip 抽圖片 + drawing XML 錨點
3. openpyxl 讀合併儲存格範圍
4. 組合產出：內嵌圖片佔位符(標註 cell) 的 .md + metadata.json
"""

import sys
import json
import shutil
import zipfile
import re
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.cell import coordinate_to_tuple
from markitdown import MarkItDown


# xlsx 內 drawing XML 的 namespace
NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def extract_images_with_anchors(xlsx_path: Path) -> list[dict]:
    """解 xlsx zip，從 drawing XML 抽取每張圖的 sheet、cell 錨點、圖片檔名"""
    results = []
    with zipfile.ZipFile(xlsx_path) as z:
        # 解析每個 sheet 對應的 drawing XML
        for sheet_name, drawing_path in _resolve_sheet_drawings(z):
            if drawing_path not in z.namelist():
                continue

            # 讀 drawing rels 取 rId→image 映射
            drawing_rels_path = _get_drawing_rels_path(drawing_path)
            rid_to_image = {}
            if drawing_rels_path in z.namelist():
                rels_xml = z.read(drawing_rels_path).decode("utf-8")
                rels_root = ET.fromstring(rels_xml)
                for rel in rels_root.findall("{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
                    rid = rel.get("Id", "")
                    target = rel.get("Target", "")
                    if "image" in target.lower() or "/media/" in target:
                        # 正規化路徑
                        if target.startswith("../"):
                            img_path = "xl/" + target[3:]
                        else:
                            img_path = target
                        rid_to_image[rid] = img_path

            # 解析 anchors
            drawing_xml = z.read(drawing_path).decode("utf-8")
            drawing_root = ET.fromstring(drawing_xml)

            for anchor in drawing_root:
                tag = anchor.tag.split("}")[-1] if "}" in anchor.tag else anchor.tag
                # twoCellAnchor / oneCellAnchor 有 <xdr:from> 錨點；absoluteAnchor 為絕對座標無 from
                if tag not in ("twoCellAnchor", "oneCellAnchor", "absoluteAnchor"):
                    continue

                # 取錨點 cell（absoluteAnchor 無 from → 標記為絕對定位，仍抽圖不丟棄）
                cell = _anchor_cell(anchor)

                # 取圖片 rId（可能在 pic/blipFill/blip 或 sp 裡）
                rid = _find_image_rid(anchor)
                if rid and rid in rid_to_image:
                    img_path = rid_to_image[rid]
                    results.append({
                        "sheet": sheet_name,
                        "cell": cell,
                        "image_path_in_zip": img_path,
                    })

    return results


def _resolve_sheet_drawings(z: zipfile.ZipFile) -> list[tuple[str, str]]:
    """回傳 [(sheet_name, drawing_path), ...]，順 worksheet rels 找到各 sheet 的 drawing XML。"""
    sheet_names = _get_sheet_names(z)
    pairs = []
    for name in z.namelist():
        m = re.match(r"xl/worksheets/_rels/sheet(\d+)\.xml\.rels", name)
        if not m:
            continue
        sheet_idx = int(m.group(1))
        rels_root = ET.fromstring(z.read(name).decode("utf-8"))
        for rel in rels_root.findall("{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
            target = rel.get("Target", "")
            if "drawing" in target.lower():
                drawing_path = target.replace("../", "xl/")
                sheet_name = sheet_names.get(sheet_idx, f"Sheet{sheet_idx}")
                pairs.append((sheet_name, drawing_path))
    return pairs


def extract_shape_texts(xlsx_path: Path) -> list[dict]:
    """抽取疊在工作表上的向量圖形（sp/cxnSp）的文字標註與其錨點 cell。
    pipeline 主抽點陣圖（blip）；callout 標註是 shape 內的 <a:t>，blip 抽不到，另抽。
    """
    NS_A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
    results = []
    with zipfile.ZipFile(xlsx_path) as z:
        for sheet_name, drawing_path in _resolve_sheet_drawings(z):
            if drawing_path not in z.namelist():
                continue
            drawing_root = ET.fromstring(z.read(drawing_path).decode("utf-8"))
            for anchor in drawing_root:
                tag = anchor.tag.split("}")[-1] if "}" in anchor.tag else anchor.tag
                if tag not in ("twoCellAnchor", "oneCellAnchor", "absoluteAnchor"):
                    continue
                # 收集 anchor 內所有文字（圖片 anchor 無 <a:t>，自然只命中 shape/連接線）
                texts = [
                    t.text.strip()
                    for t in anchor.iter(NS_A + "t")
                    if t.text and t.text.strip()
                ]
                if not texts:
                    continue
                results.append({
                    "sheet": sheet_name,
                    "cell": _anchor_cell(anchor),
                    "text": " ".join(texts),
                })
    return results


def _get_sheet_names(z: zipfile.ZipFile) -> dict[int, str]:
    """從 workbook.xml 取 sheet 順序與名稱"""
    names = {}
    wb_path = "xl/workbook.xml"
    if wb_path not in z.namelist():
        return names
    wb_xml = z.read(wb_path).decode("utf-8")
    root = ET.fromstring(wb_xml)
    ns_wb = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    sheets = root.find("s:sheets", ns_wb)
    if sheets is None:
        return names
    for i, sheet in enumerate(sheets.findall("s:sheet", ns_wb), start=1):
        names[i] = sheet.get("name", f"Sheet{i}")
    return names


def _get_drawing_rels_path(drawing_path: str) -> str:
    """drawing1.xml → xl/drawings/_rels/drawing1.xml.rels"""
    parts = drawing_path.rsplit("/", 1)
    if len(parts) == 2:
        return f"{parts[0]}/_rels/{parts[1]}.rels"
    return f"_rels/{drawing_path}.rels"


def _find_image_rid(anchor_el) -> str | None:
    """遞迴找 anchor 內的 blip embed rId"""
    for el in anchor_el.iter():
        tag = el.tag.split("}")[-1] if "}" in el.tag else el.tag
        if tag == "blip":
            return el.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
    return None


def _anchor_cell(anchor_el) -> str:
    """取 anchor 的錨點 cell。
    twoCellAnchor/oneCellAnchor 用 <xdr:from> 的 col/row（0-based）轉 A1；
    absoluteAnchor 無 from（絕對座標 EMU，無法可靠對回 cell）→ 回傳 '(absolute)'。
    """
    from_el = anchor_el.find("xdr:from", NS)
    if from_el is not None:
        col_el = from_el.find("xdr:col", NS)
        row_el = from_el.find("xdr:row", NS)
        if col_el is not None and row_el is not None:
            col = int(col_el.text)  # 0-based
            row = int(row_el.text)  # 0-based
            return get_column_letter(col + 1) + str(row + 1)
    return "(absolute)"


def get_merged_cells(xlsx_path: Path) -> list[dict]:
    """用 openpyxl 讀取所有合併儲存格範圍，並抽出左上角的標頭值。
    合併區只有左上角 cell 持有值，故 header 取 (min_row, min_col)。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    merges = []
    for ws in wb.worksheets:
        for merge_range in ws.merged_cells.ranges:
            # 取合併區左上角的值當 header（data_only 模式可直接讀值）
            min_row, min_col = merge_range.min_row, merge_range.min_col
            top_left = ws.cell(row=min_row, column=min_col).value
            header = str(top_left).strip() if top_left is not None else None
            merges.append({
                "sheet": ws.title,
                "range": str(merge_range),
                "header": header or None,  # 空字串正規化為 None
            })
    wb.close()
    return merges


def _annotate_merged_range(images: list[dict], merges: list[dict]) -> None:
    """原地標註：若圖片錨點 cell 落在某合併範圍內，填入該 range 字串。
    absoluteAnchor 的 cell='(absolute)' 無法解析座標 → 跳過（保持 None）。
    """
    # 按 sheet 建索引：[(min_col, min_row, max_col, max_row, range_str)]
    by_sheet: dict[str, list[tuple]] = {}
    for m in merges:
        min_col, min_row, max_col, max_row = range_boundaries(m["range"])
        by_sheet.setdefault(m["sheet"], []).append(
            (min_col, min_row, max_col, max_row, m["range"])
        )

    for img in images:
        try:
            row, col = coordinate_to_tuple(img["cell"])  # (row, col)
        except (ValueError, TypeError):
            continue  # 非標準 cell（如 (absolute)）
        for min_col, min_row, max_col, max_row, rng in by_sheet.get(img["sheet"], []):
            if min_row <= row <= max_row and min_col <= col <= max_col:
                img["merged_range"] = rng
                break


def _build_cell_list(ws) -> list[str]:
    """把一張 sheet 的非空 cell 輸出為 [cell] 標註清單（合併感知）。
    合併區只有左上角持值，以 range 位址呈現（如 [B38:Z38]），值不重複灌進覆蓋格——
    這是 §7.4 合併標頭「下放」的省記憶體實作：span 由 range 明示，語意等價於逐列複製。
    """
    # 合併左上角 (row,col) → range 字串；覆蓋格集合（理論無值，保險跳過）
    merge_topleft = {}
    covered = set()
    for rng in ws.merged_cells.ranges:
        merge_topleft[(rng.min_row, rng.min_col)] = str(rng)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    covered.add((r, c))

    lines = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value).strip()
            if not val:
                continue
            pos = (cell.row, cell.column)
            if pos in covered:
                continue
            addr = merge_topleft.get(pos) or f"{get_column_letter(cell.column)}{cell.row}"
            val = val.replace("\n", " / ")  # 多行壓單行，避免破壞清單
            lines.append(f"- **[{addr}]** {val}")
    return lines


def build_sheet_cell_lists(xlsx_path: Path, density_threshold: float = 0.15) -> dict[str, list[str]]:
    """對稀疏（非空格密度 < threshold）的 sheet 產生 cell 清單。
    密集 sheet 不納入（保留 MarkItDown 原表）；空 sheet 納入但清單為空（後續標記為無內容）。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    result = {}
    for ws in wb.worksheets:
        total = ws.max_row * ws.max_column
        if total <= 0:
            continue
        nonempty = sum(1 for row in ws.iter_rows() for c in row if c.value is not None)
        if nonempty / total < density_threshold:
            result[ws.title] = _build_cell_list(ws)
    wb.close()
    return result


def _replace_sparse_tables(md_text: str, cell_lists: dict[str, list[str]]) -> str:
    """把稀疏 sheet 的 NaN 表（MarkItDown 輸出）替換成 cell 清單；密集 sheet 原樣保留。"""
    lines = md_text.split("\n")
    out = []
    title = None
    body = []

    def flush():
        if title is not None and title in cell_lists:
            cl = cell_lists[title]
            out.append("")
            out.extend(cl if cl else ["（此工作表無儲存格內容）"])
            out.append("")
        else:
            out.extend(body)

    for line in lines:
        m = re.match(r"^##\s+(.+)", line)
        if m:
            flush()
            body = []
            title = m.group(1).strip()
            out.append(line)
        elif title is None:
            out.append(line)  # 首個標題前的前言，原樣保留
        else:
            body.append(line)
    flush()
    return "\n".join(out)


def convert(xlsx_path: str | Path, output_dir: str | Path = None):
    """主轉換函式"""
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        print(f"錯誤：找不到 {xlsx_path}")
        sys.exit(1)

    # 輸出目錄
    if output_dir is None:
        output_dir = Path.cwd() / "output" / xlsx_path.stem
    output_dir = Path(output_dir)

    md_dir = output_dir / "markdown"
    img_dir = output_dir / "images"
    meta_dir = output_dir / "metadata"
    src_dir = output_dir / "source"
    for d in (md_dir, img_dir, meta_dir, src_dir):
        d.mkdir(parents=True, exist_ok=True)

    # 重產前清空 images/，杜絕來源改名/增刪時殘留的孤兒舊圖
    for old in img_dir.iterdir():
        if old.is_file():
            old.unlink()

    print(f"轉換：{xlsx_path.name}")
    print(f"輸出：{output_dir}")

    # 1. 複製原始檔（來源==目的時跳過，避免 copy-to-self 的 PermissionError）
    dst_src = src_dir / xlsx_path.name
    if xlsx_path.resolve() != dst_src.resolve():
        shutil.copy2(xlsx_path, dst_src)
        print("[OK] 原始檔已複製")
    else:
        print("[OK] 來源即輸出位置，跳過複製")

    # 2. MarkItDown 抽文字/表格
    print("  抽取文字/表格（MarkItDown）...")
    mit = MarkItDown()
    result = mit.convert(str(xlsx_path))
    md_text = result.text_content
    print(f"[OK] 文字抽取完成（{len(md_text)} 字元）")

    # 3. 抽圖片 + 錨點
    print("  抽取圖片與錨點（zip/drawing XML）...")
    images = extract_images_with_anchors(xlsx_path)
    print(f"[OK] 找到 {len(images)} 張圖片")

    # 實際 dump 圖片到 images/
    extracted_images = []
    with zipfile.ZipFile(xlsx_path) as z:
        for img_info in images:
            zip_img_path = img_info["image_path_in_zip"]
            if zip_img_path in z.namelist():
                # 產生易讀檔名：sheet_cell_原始檔名（cell 可能含括號如 (absolute)，一併淨化）
                orig_name = Path(zip_img_path).name
                safe_sheet = re.sub(r'[^\w\-]', '_', img_info["sheet"])
                safe_cell = re.sub(r'[^\w\-]', '_', img_info["cell"])
                out_name = f"{safe_sheet}_{safe_cell}_{orig_name}"
                out_path = img_dir / out_name
                with z.open(zip_img_path) as src, open(out_path, "wb") as dst:
                    dst.write(src.read())
                extracted_images.append({
                    "sheet": img_info["sheet"],
                    "cell": img_info["cell"],
                    "merged_range": None,   # 由 _annotate_merged_range 補上
                    "image": f"images/{out_name}",
                    "description": None,    # 選用檢索提示（§9），預設留空
                })

    print(f"[OK] 已匯出 {len(extracted_images)} 張圖片")

    # 4. 合併儲存格
    print("  讀取合併儲存格...")
    merges = get_merged_cells(xlsx_path)
    print(f"[OK] 找到 {len(merges)} 個合併範圍")

    # 4.5 為圖片標註其錨點所在的合併範圍（若有）
    _annotate_merged_range(extracted_images, merges)

    # 4.6 抽取疊在工作表上的圖形文字標註（callout）
    print("  抽取圖形文字標註（shape/連接線）...")
    shape_texts = extract_shape_texts(xlsx_path)
    print(f"[OK] 找到 {len(shape_texts)} 筆圖形文字標註")

    # 4.7 稀疏文件型 sheet：用 cell 清單取代 NaN 灌爆的表（密度閘門）
    print("  重建稀疏工作表（cell 清單取代 NaN 表）...")
    cell_lists = build_sheet_cell_lists(xlsx_path)
    md_text = _replace_sparse_tables(md_text, cell_lists)
    print(f"[OK] 重建 {len(cell_lists)} 個稀疏工作表")

    # 5. 在 Markdown 內嵌圖片佔位符 + 圖形文字標註
    md_with_images = _inject_image_placeholders(md_text, extracted_images, shape_texts)

    # 寫出 .md
    md_file = md_dir / (xlsx_path.stem + ".md")
    md_file.write_text(md_with_images, encoding="utf-8")
    print(f"[OK] Markdown 已寫入：{md_file.name}")

    # 6. 產生 metadata.json
    metadata = {
        "source": xlsx_path.name,
        "images": extracted_images,
        "shapes": shape_texts,
        "merges": merges,
    }
    meta_file = meta_dir / "metadata.json"
    meta_file.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    print("[OK] metadata.json 已產生")

    print(f"\n完成！輸出在 {output_dir}")
    return output_dir


def _inject_image_placeholders(md_text: str, images: list[dict],
                               shapes: list[dict] = None) -> str:
    """在 Markdown 各 sheet 段落後注入圖片佔位符 + 圖形文字標註區塊"""
    shapes = shapes or []
    if not images and not shapes:
        return md_text

    # 按 sheet 分組
    img_by_sheet: dict[str, list[dict]] = {}
    for img in images:
        img_by_sheet.setdefault(img["sheet"], []).append(img)
    shp_by_sheet: dict[str, list[dict]] = {}
    for s in shapes:
        shp_by_sheet.setdefault(s["sheet"], []).append(s)

    # 在每個 sheet 標題段落後（下個標題前）插入該 sheet 的圖片+標註
    lines = md_text.split("\n")
    output_lines = []
    pending_sheet = None

    for line in lines:
        sheet_match = re.match(r"^##\s+(.+)", line)
        if sheet_match:
            if pending_sheet is not None:
                output_lines.extend(_format_sheet_extras(
                    img_by_sheet.get(pending_sheet), shp_by_sheet.get(pending_sheet)))
            pending_sheet = sheet_match.group(1).strip()
        output_lines.append(line)

    # flush 最後一個 sheet
    if pending_sheet is not None:
        output_lines.extend(_format_sheet_extras(
            img_by_sheet.get(pending_sheet), shp_by_sheet.get(pending_sheet)))

    # 沒匹配到任何標題的 sheet，內容附在最後
    matched = set()
    for line in output_lines:
        m = re.match(r"^##\s+(.+)", line)
        if m:
            matched.add(m.group(1).strip())

    unmatched_imgs = [i for s, lst in img_by_sheet.items() if s not in matched for i in lst]
    unmatched_shps = [x for s, lst in shp_by_sheet.items() if s not in matched for x in lst]
    if unmatched_imgs or unmatched_shps:
        output_lines.append("")
        output_lines.append("## 未匹配工作表內容")
        output_lines.extend(_format_sheet_extras(unmatched_imgs, unmatched_shps))

    return "\n".join(output_lines)


def _format_sheet_extras(images: list[dict], shapes: list[dict]) -> list[str]:
    """格式化單一 sheet 的圖片區 + 圖形文字標註區為 Markdown 行"""
    lines = []
    if images:
        lines += ["", "### 本工作表圖片", ""]
        for img in images:
            name = Path(img["image"]).stem
            lines.append(f'- **[{img["cell"]}]** {name} — ![{name}]({img["image"]})')
    if shapes:
        lines += ["", "### 本工作表標註（圖形文字）", ""]
        for s in shapes:
            lines.append(f'- **[{s["cell"]}]** {s["text"]}')
    if lines:
        lines.append("")
    return lines


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法：python convert.py <excel檔案路徑> [輸出目錄]")
        sys.exit(1)
    xlsx = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else None
    convert(xlsx, out)
