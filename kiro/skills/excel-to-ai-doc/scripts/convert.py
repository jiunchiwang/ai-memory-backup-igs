# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "markitdown[xlsx]>=0.0.1a2",
#   "openpyxl>=3.1.5",
# ]
# ///
"""
Excel → AI Document 轉換 pipeline

用途：把含大量示意圖的遊戲規格書(.xlsx) 轉成 AI 可理解的知識結構
輸出：markdown/ + images/ + metadata/ + source/

核心策略：
1. MarkItDown 抽文字/表格（保留原始輸出當比對基準）
2. 解 xlsx zip 抽圖片 + drawing XML 錨點
3. openpyxl 讀合併儲存格範圍 + 儲存格顏色（含佈景主題色）
4. 稀疏 sheet 重建為 cell 清單、密集 sheet 重建為合併補值表
5. 組合產出：內嵌圖片佔位符(標註 cell) 的 .md + metadata.json + 自我驗證報告

相依安裝：檔頭的 `# /// script` 是 PEP 723 inline metadata，
`uv run convert.py ...` 會自動解析並快取相依，不需先 pip install。
沒有 uv 時仍可 `pip install -r requirements.txt` 後用 python 跑。
"""

import sys
import json
import shutil
import zipfile
import re
import colorsys
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.styles.colors import COLOR_INDEX
from markitdown import MarkItDown


# xlsx 內 drawing XML 的 namespace
NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}

# 合併補值的膨脹上限：以「重複字元數」計（span 格數 × 文字長度）。
# 短標頭橫跨 90 欄無所謂；一整段文字複製十幾格才是真膨脹，改掛回指。
FULL_FILL_CHAR_CAP = 5000

# 密度閘門：非空格密度 < 此值視為「文件型」稀疏 sheet，改用 cell 清單呈現
DENSITY_THRESHOLD = 0.15

# metadata.json 內逐格樣式明細的筆數上限（超出只留圖例，避免 metadata 爆量）
STYLE_CELL_DETAIL_CAP = 20000


def extract_images_with_anchors(xlsx_path: Path) -> list[dict]:
    """解 xlsx zip，從 drawing XML 抽取每張圖的 sheet、cell 錨點、圖片檔名"""
    results = []
    with zipfile.ZipFile(xlsx_path) as z:
        # 解析每個 sheet 對應的 drawing XML
        for sheet_name, drawing_path in _resolve_sheet_drawings(z):
            if drawing_path not in z.namelist():
                continue

            # 讀 drawing rels 取 rId→image 映射
            drawing_rels_path = _get_drawing_rels_path(drawing_path)
            rid_to_image = {}
            if drawing_rels_path in z.namelist():
                rels_xml = z.read(drawing_rels_path).decode("utf-8")
                rels_root = ET.fromstring(rels_xml)
                for rel in rels_root.findall("{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
                    rid = rel.get("Id", "")
                    target = rel.get("Target", "")
                    if "image" in target.lower() or "/media/" in target:
                        # 正規化路徑
                        if target.startswith("../"):
                            img_path = "xl/" + target[3:]
                        else:
                            img_path = target
                        rid_to_image[rid] = img_path

            # 解析 anchors
            drawing_xml = z.read(drawing_path).decode("utf-8")
            drawing_root = ET.fromstring(drawing_xml)

            for anchor in drawing_root:
                tag = anchor.tag.split("}")[-1] if "}" in anchor.tag else anchor.tag
                # twoCellAnchor / oneCellAnchor 有 <xdr:from> 錨點；absoluteAnchor 為絕對座標無 from
                if tag not in ("twoCellAnchor", "oneCellAnchor", "absoluteAnchor"):
                    continue

                # 取錨點 cell（absoluteAnchor 無 from → 標記為絕對定位，仍抽圖不丟棄）
                cell = _anchor_cell(anchor)

                # 取圖片 rId（可能在 pic/blipFill/blip 或 sp 裡）
                rid = _find_image_rid(anchor)
                if rid and rid in rid_to_image:
                    img_path = rid_to_image[rid]
                    results.append({
                        "sheet": sheet_name,
                        "cell": cell,
                        "image_path_in_zip": img_path,
                    })

    return results


def _resolve_sheet_drawings(z: zipfile.ZipFile) -> list[tuple[str, str]]:
    """回傳 [(sheet_name, drawing_path), ...]，順 worksheet rels 找到各 sheet 的 drawing XML。"""
    sheet_names = _get_sheet_names(z)
    pairs = []
    for name in z.namelist():
        m = re.match(r"xl/worksheets/_rels/sheet(\d+)\.xml\.rels", name)
        if not m:
            continue
        sheet_idx = int(m.group(1))
        rels_root = ET.fromstring(z.read(name).decode("utf-8"))
        for rel in rels_root.findall("{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
            target = rel.get("Target", "")
            if "drawing" in target.lower():
                drawing_path = target.replace("../", "xl/")
                sheet_name = sheet_names.get(sheet_idx, f"Sheet{sheet_idx}")
                pairs.append((sheet_name, drawing_path))
    return pairs


def extract_shape_texts(xlsx_path: Path) -> list[dict]:
    """抽取疊在工作表上的向量圖形（sp/cxnSp）的文字標註與其錨點 cell。
    pipeline 主抽點陣圖（blip）；callout 標註是 shape 內的 <a:t>，blip 抽不到，另抽。
    """
    NS_A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
    results = []
    with zipfile.ZipFile(xlsx_path) as z:
        for sheet_name, drawing_path in _resolve_sheet_drawings(z):
            if drawing_path not in z.namelist():
                continue
            drawing_root = ET.fromstring(z.read(drawing_path).decode("utf-8"))
            for anchor in drawing_root:
                tag = anchor.tag.split("}")[-1] if "}" in anchor.tag else anchor.tag
                if tag not in ("twoCellAnchor", "oneCellAnchor", "absoluteAnchor"):
                    continue
                # 收集 anchor 內所有文字（圖片 anchor 無 <a:t>，自然只命中 shape/連接線）
                texts = [
                    t.text.strip()
                    for t in anchor.iter(NS_A + "t")
                    if t.text and t.text.strip()
                ]
                if not texts:
                    continue
                results.append({
                    "sheet": sheet_name,
                    "cell": _anchor_cell(anchor),
                    "text": " ".join(texts),
                })
    return results


def _get_sheet_names(z: zipfile.ZipFile) -> dict[int, str]:
    """從 workbook.xml 取 sheet 順序與名稱"""
    names = {}
    wb_path = "xl/workbook.xml"
    if wb_path not in z.namelist():
        return names
    wb_xml = z.read(wb_path).decode("utf-8")
    root = ET.fromstring(wb_xml)
    ns_wb = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    sheets = root.find("s:sheets", ns_wb)
    if sheets is None:
        return names
    for i, sheet in enumerate(sheets.findall("s:sheet", ns_wb), start=1):
        names[i] = sheet.get("name", f"Sheet{i}")
    return names


def _get_drawing_rels_path(drawing_path: str) -> str:
    """drawing1.xml → xl/drawings/_rels/drawing1.xml.rels"""
    parts = drawing_path.rsplit("/", 1)
    if len(parts) == 2:
        return f"{parts[0]}/_rels/{parts[1]}.rels"
    return f"_rels/{drawing_path}.rels"


def _find_image_rid(anchor_el) -> str | None:
    """遞迴找 anchor 內的 blip embed rId"""
    for el in anchor_el.iter():
        tag = el.tag.split("}")[-1] if "}" in el.tag else el.tag
        if tag == "blip":
            return el.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
    return None


def _anchor_cell(anchor_el) -> str:
    """取 anchor 的錨點 cell。
    twoCellAnchor/oneCellAnchor 用 <xdr:from> 的 col/row（0-based）轉 A1；
    absoluteAnchor 無 from（絕對座標 EMU，無法可靠對回 cell）→ 回傳 '(absolute)'。
    """
    from_el = anchor_el.find("xdr:from", NS)
    if from_el is not None:
        col_el = from_el.find("xdr:col", NS)
        row_el = from_el.find("xdr:row", NS)
        if col_el is not None and row_el is not None:
            col = int(col_el.text)  # 0-based
            row = int(row_el.text)  # 0-based
            return get_column_letter(col + 1) + str(row + 1)
    return "(absolute)"


# ---------------------------------------------------------------------------
# 顏色解析：規格書的底色/字色常帶語意（待確認 / 已砍 / 分組 / 修改標記），
# 不解出來 AI 讀到的 md 與原表語意就不同，且無從察覺。
# 三種色彩來源都要處理：rgb / indexed / theme（+tint）。
# 只認 rgb 會漏掉 Excel 調色盤上排「佈景主題色彩」——那是最常被點選的一排，
# 漏標比不做更危險（會讓人以為已覆蓋）。
# ---------------------------------------------------------------------------

# Excel 的 theme 索引順序與 theme1.xml 的 clrScheme 排列不同：
# XML 是 dk1,lt1,dk2,lt2,accent1..6,hlink,folHlink；Excel 索引前兩對互換。
_THEME_INDEX_ORDER = [
    "lt1", "dk1", "lt2", "dk2",
    "accent1", "accent2", "accent3", "accent4", "accent5", "accent6",
    "hlink", "folHlink",
]


def load_theme_colors(xlsx_path: Path) -> list[str]:
    """讀 xl/theme/theme1.xml 的 clrScheme，回傳依 Excel theme 索引排好的 RRGGBB 清單。
    讀不到（無 theme part / 格式異常）時回空清單 → theme 色一律視為無法解析。
    """
    ns_a = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            names = [n for n in z.namelist() if re.match(r"xl/theme/theme\d+\.xml", n)]
            if not names:
                return []
            root = ET.fromstring(z.read(sorted(names)[0]).decode("utf-8"))
    except (KeyError, ET.ParseError, zipfile.BadZipFile):
        return []

    scheme = root.find(f".//{ns_a}clrScheme")
    if scheme is None:
        return []

    by_name = {}
    for child in scheme:
        slot = child.tag.split("}")[-1]
        srgb = child.find(f"{ns_a}srgbClr")
        if srgb is not None and srgb.get("val"):
            by_name[slot] = srgb.get("val").upper()[-6:]
            continue
        sysclr = child.find(f"{ns_a}sysClr")
        if sysclr is not None and sysclr.get("lastClr"):
            by_name[slot] = sysclr.get("lastClr").upper()[-6:]

    return [by_name.get(slot, "") for slot in _THEME_INDEX_ORDER]


def _apply_tint(hex_rgb: str, tint: float | None) -> str:
    """套用 OOXML 的 tint（在 HLS 的亮度上調整）。tint 為 0/None 時原樣回傳。"""
    if not tint:
        return hex_rgb
    try:
        r = int(hex_rgb[0:2], 16) / 255.0
        g = int(hex_rgb[2:4], 16) / 255.0
        b = int(hex_rgb[4:6], 16) / 255.0
    except (ValueError, IndexError):
        return hex_rgb
    hue, lum, sat = colorsys.rgb_to_hls(r, g, b)
    # ECMA-376：tint<0 壓暗、tint>0 提亮（此處 HLSMAX 正規化為 1.0）
    lum = lum * (1 + tint) if tint < 0 else lum * (1 - tint) + tint
    lum = min(1.0, max(0.0, lum))
    r, g, b = colorsys.hls_to_rgb(hue, lum, sat)
    return "{:02X}{:02X}{:02X}".format(round(r * 255), round(g * 255), round(b * 255))


def resolve_color(color_obj, theme: list[str]) -> str | None:
    """把 openpyxl 的 Color 物件解成 RRGGBB；解不出來回 None。"""
    if color_obj is None:
        return None
    ctype = getattr(color_obj, "type", None)

    if ctype == "rgb":
        rgb = color_obj.rgb
        if isinstance(rgb, str) and len(rgb) >= 6:
            return _apply_tint(rgb.upper()[-6:], getattr(color_obj, "tint", 0))
        return None

    if ctype == "theme":
        idx = getattr(color_obj, "theme", None)
        if isinstance(idx, int) and 0 <= idx < len(theme) and theme[idx]:
            return _apply_tint(theme[idx], getattr(color_obj, "tint", 0))
        return None

    if ctype == "indexed":
        idx = getattr(color_obj, "indexed", None)
        if isinstance(idx, int) and 0 <= idx < len(COLOR_INDEX):
            val = COLOR_INDEX[idx]
            if isinstance(val, str) and len(val) >= 6:
                return _apply_tint(val.upper()[-6:], getattr(color_obj, "tint", 0))
        return None

    return None  # 'auto' 或未知 → 視為預設色


def cell_bg(cell, theme: list[str]) -> str | None:
    """儲存格底色；無填滿、非 solid、或純白 → None（純白等同沒上色）"""
    fill = getattr(cell, "fill", None)
    if fill is None or fill.patternType != "solid":
        return None
    hexv = resolve_color(fill.fgColor, theme)
    if hexv is None or hexv == "FFFFFF":
        return None
    return hexv


def cell_font_color(cell, theme: list[str]) -> str | None:
    """儲存格字色；預設黑 → None"""
    font = getattr(cell, "font", None)
    if font is None:
        return None
    hexv = resolve_color(font.color, theme)
    if hexv is None or hexv == "000000":
        return None
    return hexv


class StyleRegistry:
    """把 (底色, 字色) 組合登錄成短代號 c1/c2...，md 內文只掛代號、hex 集中放圖例。
    逐格重複 hex 在大量上色的規格書會嚴重膨脹 token；代號+圖例同資訊量但便宜得多，
    也讓 AI 一眼看出「有幾種色彩分組」而不是被一堆 hex 淹沒。
    """

    def __init__(self):
        self._count: dict[tuple, int] = {}
        self._sample: dict[tuple, str] = {}
        self._id: dict[tuple, str] = {}
        self.cells: list[dict] = []
        self.detail_truncated = 0

    def observe(self, sheet: str, addr: str, bg: str | None, font: str | None) -> None:
        if bg is None and font is None:
            return
        key = (bg, font)
        self._count[key] = self._count.get(key, 0) + 1
        self._sample.setdefault(key, f"{sheet}!{addr}")

    def finalize(self) -> None:
        """依出現次數由多到少配代號（常見色分組拿到短號，人與 AI 都好對照）"""
        ordered = sorted(self._count.items(), key=lambda kv: (-kv[1], kv[0][0] or "", kv[0][1] or ""))
        for i, (key, _) in enumerate(ordered, start=1):
            self._id[key] = f"c{i}"

    def tag(self, sheet: str, addr: str, bg: str | None, font: str | None) -> str:
        """回傳要接在文字後的 ` [c3]`；無色時回空字串。順便記錄逐格明細。"""
        if bg is None and font is None:
            return ""
        sid = self._id.get((bg, font))
        if sid is None:
            return ""
        if len(self.cells) < STYLE_CELL_DETAIL_CAP:
            self.cells.append({"sheet": sheet, "cell": addr, "style": sid})
        else:
            self.detail_truncated += 1
        return f" [{sid}]"

    def legend(self) -> list[dict]:
        out = []
        for key, sid in sorted(self._id.items(), key=lambda kv: int(kv[1][1:])):
            bg, font = key
            out.append({
                "id": sid,
                "bg": f"#{bg}" if bg else None,
                "font": f"#{font}" if font else None,
                "count": self._count[key],
                "sample": self._sample[key],
            })
        return out

    def legend_lines(self) -> list[str]:
        leg = self.legend()
        if not leg:
            return []
        lines = ["", "## 顏色圖例", "",
                 "> 內文儲存格後的 `[c1]` 代號對應下表。規格書的底色/字色通常帶語意"
                 "（分組、待確認、已砍、修改標記），判讀內容時一併參考。", "",
                 "| 代號 | 底色 | 字色 | 出現次數 | 範例位置 |",
                 "| --- | --- | --- | --- | --- |"]
        for e in leg:
            lines.append(
                f'| {e["id"]} | {e["bg"] or "-"} | {e["font"] or "-"} | {e["count"]} | {e["sample"]} |'
            )
        lines.append("")
        return lines


def scan_styles(wb, theme: list[str]) -> StyleRegistry:
    """全 workbook 掃一遍顏色，先建圖例再渲染（代號要照全域出現次數排序）。"""
    reg = StyleRegistry()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                # 只登錄有文字的格：無內容的色塊多半是排版分隔色帶，掛 tag 只會
                # 產生整列 [c5] [c5]... 的純噪音（實測一張多國語言表就吃掉 27 格）
                if not _has_data(cell.value):
                    continue
                bg = cell_bg(cell, theme)
                fc = cell_font_color(cell, theme)
                if bg is None and fc is None:
                    continue
                addr = f"{get_column_letter(cell.column)}{cell.row}"
                reg.observe(ws.title, addr, bg, fc)
    reg.finalize()
    return reg


def load_workbook_with_formulas(xlsx_path: Path) -> tuple[object, dict[str, int]]:
    """載入活頁簿，並把「有公式但沒有 cached 結果」的格補回公式字串。

    openpyxl 的 data_only=True 讀到的是 Excel 上次存檔時算好的快取值，不是公式
    本身。由 openpyxl 之類的產生器寫出、或存檔前未重算的檔案沒有快取，那些格會
    讀成 None——一整張以公式為主的工作表會被判成空白而靜默消失，下游卻把這份
    輸出當 ground truth。回填公式字串至少讓內容有出口；同時回傳每張表的回填
    筆數，讓驗證層據以警告（公式字串不等於計算結果，使用者仍該回 Excel 重存）。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        wbf = openpyxl.load_workbook(xlsx_path, data_only=False)
    except Exception:
        return wb, {}  # 讀不到公式視圖就維持原行為，不讓它擋住整條轉換

    counts: dict[str, int] = {}
    try:
        for ws in wb.worksheets:
            if ws.title not in wbf.sheetnames:
                continue
            n = 0
            for row in wbf[ws.title].iter_rows():
                for fcell in row:
                    if not _has_data(fcell.value):
                        continue
                    target = ws.cell(row=fcell.row, column=fcell.column)
                    if _has_data(target.value):
                        continue
                    try:
                        target.value = fcell.value
                    except AttributeError:
                        continue  # MergedCell 唯讀，覆蓋格本來就不該持值
                    n += 1
            if n:
                counts[ws.title] = n
    finally:
        wbf.close()
    return wb, counts


def get_merged_cells(wb) -> list[dict]:
    """讀取所有合併儲存格範圍，並抽出左上角的標頭值。
    合併區只有左上角 cell 持有值，故 header 取 (min_row, min_col)。
    """
    merges = []
    for ws in wb.worksheets:
        for merge_range in ws.merged_cells.ranges:
            # 取合併區左上角的值當 header（data_only 模式可直接讀值）
            min_row, min_col = merge_range.min_row, merge_range.min_col
            top_left = ws.cell(row=min_row, column=min_col).value
            header = str(top_left).strip() if top_left is not None else None
            merges.append({
                "sheet": ws.title,
                "range": str(merge_range),
                "header": header or None,  # 空字串正規化為 None
            })
    return merges


def _annotate_merged_range(images: list[dict], merges: list[dict]) -> None:
    """原地標註：若圖片錨點 cell 落在某合併範圍內，填入該 range 字串。
    absoluteAnchor 的 cell='(absolute)' 無法解析座標 → 跳過（保持 None）。
    """
    # 按 sheet 建索引：[(min_col, min_row, max_col, max_row, range_str)]
    by_sheet: dict[str, list[tuple]] = {}
    for m in merges:
        min_col, min_row, max_col, max_row = range_boundaries(m["range"])
        by_sheet.setdefault(m["sheet"], []).append(
            (min_col, min_row, max_col, max_row, m["range"])
        )

    for img in images:
        try:
            row, col = coordinate_to_tuple(img["cell"])  # (row, col)
        except (ValueError, TypeError):
            continue  # 非標準 cell（如 (absolute)）
        for min_col, min_row, max_col, max_row, rng in by_sheet.get(img["sheet"], []):
            if min_row <= row <= max_row and min_col <= col <= max_col:
                img["merged_range"] = rng
                break


# ---------------------------------------------------------------------------
# Sheet 內容重建：稀疏 → cell 清單；密集 → 合併補值表
# ---------------------------------------------------------------------------


def _has_data(v) -> bool:
    """None 與純空白都不算資料——手動排版殘留的一個空格不該把邊界撐大一整欄。"""
    if v is None:
        return False
    if isinstance(v, str) and v.strip() == "":
        return False
    return True


def _cell_text(v) -> str:
    """轉成可放進 GFM 表格的一行字：多行併 <br>、跳脫管線符號。"""
    if v is None:
        return ""
    s = str(v)
    lines = [ln.strip() for ln in s.splitlines() if ln.strip()]
    return "<br>".join(lines).replace("|", "\\|")


def _build_cell_list(ws, theme: list[str], reg: StyleRegistry) -> tuple[list[str], int]:
    """把一張 sheet 的非空 cell 輸出為 [cell] 標註清單（合併感知）。
    合併區只有左上角持值，以 range 位址呈現（如 [B38:Z38]），值不重複灌進覆蓋格——
    這是 §7.4 合併標頭「下放」的省記憶體實作：span 由 range 明示，語意等價於逐列複製。
    回傳 (清單行, 呈現的 cell 數)。
    """
    # 合併左上角 (row,col) → range 字串；覆蓋格集合（理論無值，保險跳過）
    merge_topleft = {}
    covered = set()
    for rng in ws.merged_cells.ranges:
        merge_topleft[(rng.min_row, rng.min_col)] = str(rng)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    covered.add((r, c))

    lines = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value).strip()
            if not val:
                continue
            pos = (cell.row, cell.column)
            if pos in covered:
                continue
            plain_addr = f"{get_column_letter(cell.column)}{cell.row}"
            addr = merge_topleft.get(pos) or plain_addr
            val = val.replace("\n", " / ")  # 多行壓單行，避免破壞清單
            tag = reg.tag(ws.title, plain_addr, cell_bg(cell, theme), cell_font_color(cell, theme))
            lines.append(f"- **[{addr}]** {val}{tag}")
    return lines, len(lines)


def _sheet_bounds(ws) -> tuple[int, int, int, int] | None:
    """有資料的儲存格 + 合併範圍共同決定的邊界框；全空回 None。
    合併範圍要納入：合併原點在邊界上但 span 更遠時，只看資料格會把 span 切掉。
    """
    min_row, max_row = ws.max_row + 1, 0
    min_col, max_col = ws.max_column + 1, 0
    for row in ws.iter_rows():
        for cell in row:
            if _has_data(cell.value):
                min_row, max_row = min(min_row, cell.row), max(max_row, cell.row)
                min_col, max_col = min(min_col, cell.column), max(max_col, cell.column)
    for m in ws.merged_cells.ranges:
        if not _has_data(ws.cell(row=m.min_row, column=m.min_col).value):
            continue
        min_row, max_row = min(min_row, m.min_row), max(max_row, m.max_row)
        min_col, max_col = min(min_col, m.min_col), max(max_col, m.max_col)
    if max_row < min_row or max_col < min_col:
        return None
    return min_row, max_row, min_col, max_col


def _build_dense_table(ws, theme: list[str], reg: StyleRegistry) -> tuple[list[str], dict]:
    """密集 sheet 重建成 GFM 表：合併值補滿整個 span、顏色掛代號。
    GFM 表無 colspan/rowspan，合併資訊只能靠補值表達；不補的話覆蓋格全是空白，
    AI 會把「屬於上一個合併標頭的列」讀成無標頭的孤立列。
    """
    bounds = _sheet_bounds(ws)
    if bounds is None:
        return [], {"rows": 0, "cols": 0, "merged_filled": 0, "merged_capped": 0}
    min_row, max_row, min_col, max_col = bounds

    origin_of: dict[tuple, tuple] = {}
    span_chars: dict[tuple, int] = {}
    for m in ws.merged_cells.ranges:
        origin = (m.min_row, m.min_col)
        top_left = ws.cell(row=m.min_row, column=m.min_col).value
        if not _has_data(top_left):
            continue
        covered = 0
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                if (r, c) != origin:
                    origin_of[(r, c)] = origin
                    covered += 1
        span_chars[origin] = covered * len(str(top_left))

    n_filled = n_capped = 0
    grid = []
    for r in range(min_row, max_row + 1):
        row_out = []
        for c in range(min_col, max_col + 1):
            origin = origin_of.get((r, c))
            if origin is not None:
                oc = ws.cell(row=origin[0], column=origin[1])
                if span_chars[origin] > FULL_FILL_CHAR_CAP:
                    # 整段文字複製滿 span 會膨脹，改掛回指（語意不失、體積可控）
                    text = f"(→ 合併自 {get_column_letter(origin[1])}{origin[0]})"
                    n_capped += 1
                else:
                    text = _cell_text(oc.value)
                    n_filled += 1
                style_cell = oc
            else:
                style_cell = ws.cell(row=r, column=c)
                text = _cell_text(style_cell.value)
            if text:  # 空格不掛顏色 tag，與 scan_styles 的登錄條件保持一致
                addr = f"{get_column_letter(c)}{r}"
                text += reg.tag(ws.title, addr,
                                cell_bg(style_cell, theme), cell_font_color(style_cell, theme))
            row_out.append(text.strip())
        grid.append(row_out)

    width = max_col - min_col + 1
    lines = ["| " + " | ".join(grid[0]) + " |",
             "| " + " | ".join(["---"] * width) + " |"]
    for row in grid[1:]:
        lines.append("| " + " | ".join(row) + " |")

    stats = {"rows": len(grid), "cols": width,
             "merged_filled": n_filled, "merged_capped": n_capped}
    return lines, stats


def build_sheet_bodies(wb, theme: list[str], reg: StyleRegistry,
                       density_threshold: float = DENSITY_THRESHOLD
                       ) -> tuple[dict[str, list[str]], list[dict]]:
    """為每個 sheet 產生要取代 MarkItDown 輸出的內容。
    稀疏（非空格密度 < threshold）→ cell 清單；其餘 → 合併補值表。
    回傳 ({sheet: 行}, per-sheet stats)。
    """
    bodies: dict[str, list[str]] = {}
    stats: list[dict] = []
    for ws in wb.worksheets:
        total = ws.max_row * ws.max_column
        nonempty = sum(1 for row in ws.iter_rows() for c in row if _has_data(c.value))
        density = (nonempty / total) if total > 0 else 0.0

        if total <= 0 or nonempty == 0:
            bodies[ws.title] = []
            stats.append({"sheet": ws.title, "mode": "empty", "nonempty_cells": 0,
                          "rows": 0, "cols": 0, "entries": 0,
                          "merged_filled": 0, "merged_capped": 0})
            continue

        if density < density_threshold:
            lines, entries = _build_cell_list(ws, theme, reg)
            bodies[ws.title] = lines
            stats.append({"sheet": ws.title, "mode": "sparse", "nonempty_cells": nonempty,
                          "rows": len(lines), "cols": 1, "entries": entries,
                          "merged_filled": 0, "merged_capped": 0})
        else:
            lines, st = _build_dense_table(ws, theme, reg)
            bodies[ws.title] = lines
            stats.append({"sheet": ws.title, "mode": "dense", "nonempty_cells": nonempty,
                          "rows": st["rows"], "cols": st["cols"],
                          "entries": st["rows"] * st["cols"],
                          "merged_filled": st["merged_filled"],
                          "merged_capped": st["merged_capped"]})
    return bodies, stats


def _replace_sheet_bodies(md_text: str, bodies: dict[str, list[str]]) -> str:
    """用重建內容取代 MarkItDown 各 sheet 段落；沒對應到的 sheet 補在最後。"""
    lines = md_text.split("\n")
    out: list[str] = []
    title = None
    body: list[str] = []
    seen: set[str] = set()

    def flush():
        if title is not None and title in bodies:
            seen.add(title)
            content = bodies[title]
            out.append("")
            out.extend(content if content else ["（此工作表無儲存格內容）"])
            out.append("")
        else:
            out.extend(body)

    for line in lines:
        m = re.match(r"^##\s+(.+)", line)
        if m:
            flush()
            body = []
            title = m.group(1).strip()
            out.append(line)
        elif title is None:
            out.append(line)  # 首個標題前的前言，原樣保留
        else:
            body.append(line)
    flush()

    # MarkItDown 沒吐出標題的 sheet（少見）不能就這樣消失
    for name, content in bodies.items():
        if name in seen:
            continue
        out += ["", f"## {name}", ""]
        out.extend(content if content else ["（此工作表無儲存格內容）"])
        out.append("")

    return "\n".join(out)


# ---------------------------------------------------------------------------
# 自我驗證：輸出不能只印「成功」，要拿兩份獨立推導互相對照才算驗到東西
# ---------------------------------------------------------------------------

_IMG_EXT = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".emf", ".wmf", ".tif", ".tiff", ".svg", ".webp"}


def _count_pipes(line: str) -> int:
    """數未跳脫的 `|`（內容裡的 `\\|` 不算欄位分隔）"""
    return len(re.findall(r"(?<!\\)\|", line))


def _parse_md_shapes(md_text: str) -> dict[str, tuple[int, int]]:
    """從 md 解出每個 sheet 的 (行數, 欄數)，供兩版對照。"""
    parts = re.split(r"^##\s+(.+)$", md_text, flags=re.MULTILINE)
    shapes: dict[str, tuple[int, int]] = {}
    for i in range(1, len(parts), 2):
        name = parts[i].strip()
        body = [ln for ln in parts[i + 1].split("\n") if ln.strip()]
        table = [ln for ln in body if ln.lstrip().startswith("|")]
        if not table:
            shapes[name] = (len(body), 0)
            continue
        n_cols = max(_count_pipes(ln) - 1 for ln in table)
        n_rows = len(table) - 1  # 扣掉 GFM 分隔列
        shapes[name] = (n_rows, n_cols)
    return shapes


def build_validation(xlsx_path: Path, raw_md: str, final_md: str,
                     sheet_stats: list[dict], extracted_images: list[dict],
                     reg: StyleRegistry, formula_fallbacks: dict[str, int]
                     ) -> tuple[str, dict]:
    """比對四件事：圖片有無漏、sheet 形狀有無縮水、非空格有無出口、公式有無快取。"""
    lines: list[str] = []
    summary: dict = {}

    # 1) 圖片：xl/media 內的檔案數 vs 實際抽出數。差額 = 沒有 anchor 的孤兒圖，
    #    這是 SKILL.md 原本要人工「自行確認 xl/media」的那一步，改成機械檢查。
    with zipfile.ZipFile(xlsx_path) as z:
        media = [n for n in z.namelist()
                 if n.startswith("xl/media/") and Path(n).suffix.lower() in _IMG_EXT]
    # 用 zip 內原始路徑比對，不要從輸出檔名反推——sheet 名淨化後含多個底線，切不回來
    referenced = {img.get("source_in_zip") for img in extracted_images}
    orphans = [n for n in media if n not in referenced]
    summary["media_files"] = len(media)
    summary["images_extracted"] = len(extracted_images)
    summary["orphan_media"] = orphans
    lines.append(f"[圖片] xl/media 檔案數={len(media)}，已錨定抽出={len(extracted_images)}")
    if orphans:
        lines.append(f"  ⚠ {len(orphans)} 張圖在 xl/media 但無 drawing 錨點（可能是背景圖/被刪 anchor 的殘留）：")
        for n in orphans[:10]:
            lines.append(f"    - {n}")
        if len(orphans) > 10:
            lines.append(f"    ...另 {len(orphans) - 10} 張")
    else:
        lines.append("  所有 media 圖片皆已錨定抽出。")

    # 2) sheet 形狀：重建版 vs MarkItDown 原版
    raw_shapes = _parse_md_shapes(raw_md)
    final_shapes = _parse_md_shapes(final_md)
    missing = [n for n in raw_shapes if n not in final_shapes]
    shrunk = []
    for name, (r_rows, r_cols) in raw_shapes.items():
        if name not in final_shapes:
            continue
        f_rows, f_cols = final_shapes[name]
        mode = next((s["mode"] for s in sheet_stats if s["sheet"] == name), None)
        if mode != "dense":
            # 稀疏表已改成清單、空表只剩一句說明、非 sheet 標題（如「未匹配工作表內容」）
            # 都沒有可比的表格形狀，硬比只會製造假警報
            continue
        if f_rows < r_rows or f_cols < r_cols:
            shrunk.append((name, (r_rows, r_cols), (f_rows, f_cols)))
    summary["sheets_missing"] = missing
    summary["sheets_shrunk"] = [{"sheet": n, "markitdown": a, "rebuilt": b} for n, a, b in shrunk]
    # 最終 md 除了工作表還有「未匹配工作表內容」等附加段落，計數只認真正的工作表
    real_sheets = {s["sheet"] for s in sheet_stats}
    lines.append(f"[形狀] 活頁簿工作表={len(real_sheets)}，MarkItDown 段落={len(raw_shapes)}，"
                 f"最終輸出含工作表={len(real_sheets & set(final_shapes))}")
    if missing:
        lines.append(f"  ⚠ 最終輸出缺少的工作表：{missing}")
    if shrunk:
        lines.append(f"  ⚠ {len(shrunk)} 個密集表比 MarkItDown 版小（可能有資料流失，需查）：")
        for n, a, b in shrunk:
            lines.append(f"    - {n}: MarkItDown={a} vs 重建={b}")
    if not missing and not shrunk:
        lines.append("  無工作表遺失，密集表皆未縮水（重建版 ≥ MarkItDown 版）。")

    # 3) 資料出口：每個 sheet 的非空格數 vs 輸出呈現的條目
    lines.append("[覆蓋] 各工作表非空格數 → 輸出條目")
    zero_out = []
    for s in sheet_stats:
        if s["mode"] == "empty":
            continue
        lines.append(f"  - {s['sheet']} [{s['mode']}] 非空格={s['nonempty_cells']}，"
                     f"輸出={s['rows']}列×{s['cols']}欄"
                     + (f"，合併補值={s['merged_filled']} 截斷回指={s['merged_capped']}"
                        if s["mode"] == "dense" else ""))
        if s["nonempty_cells"] > 0 and s["rows"] == 0:
            zero_out.append(s["sheet"])
    if zero_out:
        lines.append(f"  ⚠ 有內容卻輸出為空的工作表：{zero_out}")
    summary["sheets_empty_output"] = zero_out

    # 4) 公式快取。data_only 讀的是 Excel 存檔時算好的值；沒快取的公式格會是 None，
    #    整張以公式為主的表會被判空而消失。已回填公式字串，但那不是計算結果。
    total_fb = sum(formula_fallbacks.values())
    summary["formula_fallback_cells"] = total_fb
    summary["formula_fallback_sheets"] = dict(formula_fallbacks)
    if total_fb:
        lines.append(f"[公式] ⚠ {total_fb} 格有公式但無 cached 結果，已回填公式字串（不是計算結果）：")
        for name, n in formula_fallbacks.items():
            lines.append(f"    - {name}: {n} 格")
        lines.append("    請在 Excel 開啟後重新存檔以寫入快取值，再轉一次；否則這些格的數值缺席。")
    else:
        lines.append("[公式] 無公式快取缺漏。")

    # 5) 顏色
    leg = reg.legend()
    summary["style_variants"] = len(leg)
    summary["styled_cells"] = sum(e["count"] for e in leg)
    lines.append(f"[顏色] 相異樣式組合={len(leg)}，上色儲存格={summary['styled_cells']}")
    if reg.detail_truncated:
        lines.append(f"  ⚠ 逐格明細超過 {STYLE_CELL_DETAIL_CAP} 筆上限，metadata 少收 {reg.detail_truncated} 筆（圖例不受影響）")

    ok = not (missing or shrunk or zero_out or total_fb)
    lines.insert(0, "整體：通過（無遺失/縮水/空輸出/公式缺快取）" if ok
                 else "整體：有需要人工確認的項目，見下方 ⚠")
    summary["ok"] = ok
    return "\n".join(lines), summary


def convert(xlsx_path: str | Path, output_dir: str | Path = None):
    """主轉換函式"""
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        print(f"錯誤：找不到 {xlsx_path}")
        sys.exit(1)

    # 輸出目錄
    if output_dir is None:
        output_dir = Path.cwd() / "output" / xlsx_path.stem
    output_dir = Path(output_dir)

    md_dir = output_dir / "markdown"
    img_dir = output_dir / "images"
    meta_dir = output_dir / "metadata"
    src_dir = output_dir / "source"
    for d in (md_dir, img_dir, meta_dir, src_dir):
        d.mkdir(parents=True, exist_ok=True)

    # 重產前清空 images/，杜絕來源改名/增刪時殘留的孤兒舊圖
    for old in img_dir.iterdir():
        if old.is_file():
            old.unlink()

    print(f"轉換：{xlsx_path.name}")
    print(f"輸出：{output_dir}")

    # 1. 複製原始檔（來源==目的時跳過，避免 copy-to-self 的 PermissionError）
    dst_src = src_dir / xlsx_path.name
    if xlsx_path.resolve() != dst_src.resolve():
        shutil.copy2(xlsx_path, dst_src)
        print("[OK] 原始檔已複製")
    else:
        print("[OK] 來源即輸出位置，跳過複製")

    # 2. MarkItDown 抽文字/表格
    print("  抽取文字/表格（MarkItDown）...")
    mit = MarkItDown()
    result = mit.convert(str(xlsx_path))
    raw_md = result.text_content
    raw_md_file = md_dir / (xlsx_path.stem + "_markitdown_raw.md")
    raw_md_file.write_text(raw_md, encoding="utf-8")
    print(f"[OK] 文字抽取完成（{len(raw_md)} 字元），原始版留底：{raw_md_file.name}")

    # 3. 抽圖片 + 錨點
    print("  抽取圖片與錨點（zip/drawing XML）...")
    images = extract_images_with_anchors(xlsx_path)
    print(f"[OK] 找到 {len(images)} 張圖片")

    # 實際 dump 圖片到 images/
    extracted_images = []
    with zipfile.ZipFile(xlsx_path) as z:
        for img_info in images:
            zip_img_path = img_info["image_path_in_zip"]
            if zip_img_path in z.namelist():
                # 產生易讀檔名：sheet_cell_原始檔名（cell 可能含括號如 (absolute)，一併淨化）
                orig_name = Path(zip_img_path).name
                safe_sheet = re.sub(r'[^\w\-]', '_', img_info["sheet"])
                safe_cell = re.sub(r'[^\w\-]', '_', img_info["cell"])
                out_name = f"{safe_sheet}_{safe_cell}_{orig_name}"
                out_path = img_dir / out_name
                with z.open(zip_img_path) as src, open(out_path, "wb") as dst:
                    dst.write(src.read())
                extracted_images.append({
                    "sheet": img_info["sheet"],
                    "cell": img_info["cell"],
                    "merged_range": None,   # 由 _annotate_merged_range 補上
                    "image": f"images/{out_name}",
                    "source_in_zip": zip_img_path,  # 孤兒圖驗證的比對鍵
                    "description": None,    # 選用檢索提示（§9），預設留空
                })

    print(f"[OK] 已匯出 {len(extracted_images)} 張圖片")

    # 4. 載入活頁簿（整條流程只載這一次）+ 合併儲存格
    print("  載入活頁簿（公式無快取時回填公式字串）...")
    wb, formula_fallbacks = load_workbook_with_formulas(xlsx_path)
    if formula_fallbacks:
        print(f"[!!] {sum(formula_fallbacks.values())} 格公式無 cached 結果，已回填公式字串")
    merges = get_merged_cells(wb)
    print(f"[OK] 找到 {len(merges)} 個合併範圍")

    # 4.5 為圖片標註其錨點所在的合併範圍（若有）
    _annotate_merged_range(extracted_images, merges)

    # 4.6 抽取疊在工作表上的圖形文字標註（callout）
    print("  抽取圖形文字標註（shape/連接線）...")
    shape_texts = extract_shape_texts(xlsx_path)
    print(f"[OK] 找到 {len(shape_texts)} 筆圖形文字標註")

    # 4.7 顏色掃描（含佈景主題色解析）→ 建圖例
    print("  掃描儲存格顏色（含 theme/indexed 解析）...")
    theme = load_theme_colors(xlsx_path)
    reg = scan_styles(wb, theme)
    print(f"[OK] 相異樣式組合 {len(reg.legend())} 種"
          + ("" if theme else "（⚠ 讀不到 theme1.xml，佈景主題色將無法解析）"))

    # 4.8 重建各 sheet：稀疏→cell 清單、密集→合併補值表
    print("  重建工作表內容（稀疏=cell 清單 / 密集=合併補值表）...")
    bodies, sheet_stats = build_sheet_bodies(wb, theme, reg)
    md_text = _replace_sheet_bodies(raw_md, bodies)
    n_sparse = sum(1 for s in sheet_stats if s["mode"] == "sparse")
    n_dense = sum(1 for s in sheet_stats if s["mode"] == "dense")
    print(f"[OK] 稀疏 {n_sparse} 個、密集 {n_dense} 個")
    wb.close()

    # 5. 在 Markdown 內嵌圖片佔位符 + 圖形文字標註 + 顏色圖例
    md_with_images = _inject_image_placeholders(md_text, extracted_images, shape_texts)
    legend_lines = reg.legend_lines()
    if legend_lines:
        md_with_images = md_with_images.rstrip("\n") + "\n" + "\n".join(legend_lines)

    # 寫出 .md
    md_file = md_dir / (xlsx_path.stem + ".md")
    md_file.write_text(md_with_images, encoding="utf-8")
    print(f"[OK] Markdown 已寫入：{md_file.name}")

    # 6. 產生 metadata.json
    metadata = {
        "source": xlsx_path.name,
        "images": extracted_images,
        "shapes": shape_texts,
        "merges": merges,
        "style_legend": reg.legend(),
        "styled_cells": reg.cells,
    }
    meta_file = meta_dir / "metadata.json"
    meta_file.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    print("[OK] metadata.json 已產生")

    # 7. 自我驗證 + 統計
    print("  自我驗證（圖片/形狀/覆蓋）...")
    report, summary = build_validation(xlsx_path, raw_md, md_with_images,
                                       sheet_stats, extracted_images, reg,
                                       formula_fallbacks)
    (meta_dir / "validation.txt").write_text(report, encoding="utf-8")
    (meta_dir / "stats.json").write_text(
        json.dumps({"summary": summary, "sheets": sheet_stats},
                   ensure_ascii=False, indent=2), encoding="utf-8")
    print("      " + report.replace("\n", "\n      "))

    print(f"\n完成！輸出在 {output_dir}")
    return output_dir


def _inject_image_placeholders(md_text: str, images: list[dict],
                               shapes: list[dict] = None) -> str:
    """在 Markdown 各 sheet 段落後注入圖片佔位符 + 圖形文字標註區塊"""
    shapes = shapes or []
    if not images and not shapes:
        return md_text

    # 按 sheet 分組
    img_by_sheet: dict[str, list[dict]] = {}
    for img in images:
        img_by_sheet.setdefault(img["sheet"], []).append(img)
    shp_by_sheet: dict[str, list[dict]] = {}
    for s in shapes:
        shp_by_sheet.setdefault(s["sheet"], []).append(s)

    # 在每個 sheet 標題段落後（下個標題前）插入該 sheet 的圖片+標註
    lines = md_text.split("\n")
    output_lines = []
    pending_sheet = None

    for line in lines:
        sheet_match = re.match(r"^##\s+(.+)", line)
        if sheet_match:
            if pending_sheet is not None:
                output_lines.extend(_format_sheet_extras(
                    img_by_sheet.get(pending_sheet), shp_by_sheet.get(pending_sheet)))
            pending_sheet = sheet_match.group(1).strip()
        output_lines.append(line)

    # flush 最後一個 sheet
    if pending_sheet is not None:
        output_lines.extend(_format_sheet_extras(
            img_by_sheet.get(pending_sheet), shp_by_sheet.get(pending_sheet)))

    # 沒匹配到任何標題的 sheet，內容附在最後
    matched = set()
    for line in output_lines:
        m = re.match(r"^##\s+(.+)", line)
        if m:
            matched.add(m.group(1).strip())

    unmatched_imgs = [i for s, lst in img_by_sheet.items() if s not in matched for i in lst]
    unmatched_shps = [x for s, lst in shp_by_sheet.items() if s not in matched for x in lst]
    if unmatched_imgs or unmatched_shps:
        output_lines.append("")
        output_lines.append("## 未匹配工作表內容")
        output_lines.extend(_format_sheet_extras(unmatched_imgs, unmatched_shps))

    return "\n".join(output_lines)


def _format_sheet_extras(images: list[dict], shapes: list[dict]) -> list[str]:
    """格式化單一 sheet 的圖片區 + 圖形文字標註區為 Markdown 行"""
    lines = []
    if images:
        lines += ["", "### 本工作表圖片", ""]
        for img in images:
            name = Path(img["image"]).stem
            # metadata 的 image 是「相對輸出根目錄」，但 md 本身在 markdown/ 底下，
            # 直接照抄會解析成 markdown/images/... 這個不存在的路徑，圖全部失效
            lines.append(f'- **[{img["cell"]}]** {name} — ![{name}](../{img["image"]})')
    if shapes:
        lines += ["", "### 本工作表標註（圖形文字）", ""]
        for s in shapes:
            lines.append(f'- **[{s["cell"]}]** {s["text"]}')
    if lines:
        lines.append("")
    return lines


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法：python convert.py <excel檔案路徑> [輸出目錄]")
        sys.exit(1)
    xlsx = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else None
    convert(xlsx, out)
