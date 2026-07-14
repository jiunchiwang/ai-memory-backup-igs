"""
uk-slot-spec-adapter / spec_adapter.py
======================================

Convert a UK slot specification xlsx (the kind produced by the design team, e.g.
"三幣瑞龍專案規格文件.xlsx") into a Markdown-formatted `Game_Spec.md` suitable
as the stage-1 input for the `slot-game-codegen-skill` Kiro skill.

Scope: xlsx only in v1. docx/pdf will be added later.

Design principles:
  * Label-based lookup (find "遊戲名稱(中文)" → take the cell to the right of it),
    not hard-coded row numbers. This makes the adapter robust to minor xlsx edits.
  * Section-based parsing inside the `基本規格` sheet. Each section is delimited
    by a "■xxx" marker row. We walk rows after the marker until we hit the next
    marker or a long run of empty rows.
  * Never fail silently: any field we cannot extract is emitted as a ⚠ warning in
    the output Markdown (and also in the `Parse Warnings` section at the top).

CLI:
    py spec_adapter.py <input.xlsx> <output.md>

Returns:
    exit code 0 on success (even if there are warnings),
    exit code 2 on usage error,
    exit code 3 on unrecoverable failure (file missing / not xlsx / no basic
    sheet, etc.).
"""
from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Data classes (Game_Spec intermediate representation)
# ---------------------------------------------------------------------------


@dataclass
class ProjectInfo:
    name_zh: str = ""
    name_en: str = ""
    game_id: str = ""
    short_name: str = ""


@dataclass
class SymbolRow:
    index: int
    enum_name: str          # canonical enum member name (alias if in whitelist, else Symbol_NN)
    raw_enum_name: str      # original '系統命名' column value, always e.g. "Symbol_00"
    display: str            # original 樣式 column, e.g. "9", "10", "J", "盆栽"
    category: str           # "low"/"high"/"wild"/"scatter"/"special"  (server-only → "special" + server_only=True)
    server_only: bool = False  # True when 基本規格 sheet col B = "Server用"


@dataclass
class CheatKeyRow:
    effect: str   # e.g. "必進FG(猛虎)"
    code: str     # e.g. "1012"
    note: str


@dataclass
class AudioClip:
    section: str        # BGM / MG通用 / FG相關 / 補充
    item_zh: str        # "停輪音效"
    filename: str       # "reel_stop"
    note: str


@dataclass
class ParsedSpec:
    project: ProjectInfo = field(default_factory=ProjectInfo)
    pay_mode: str = ""         # LINE_PAY / SCATTER_PAY / UNKNOWN
    col: int | None = None
    row: int | None = None
    board_layout: str = ""     # e.g. "5x4x4x4x4x5" (per-column row counts) or "" if uniform
    symbol_notes: str = ""     # e.g. "6小、5大、3SCATTER、1WILD"
    symbols: list[SymbolRow] = field(default_factory=list)
    cheat_keys: list[CheatKeyRow] = field(default_factory=list)
    audio_clips: list[AudioClip] = field(default_factory=list)
    has_free_game: bool = False
    has_jackpot: bool = False
    has_buy_bonus: bool = False
    custom_features: list[str] = field(default_factory=list)  # non-standard features detected from spec
    performance_hints: list[dict] = field(default_factory=list)  # [{'sheet': str, 'kind': str, 'summary': str, 'row_count': int}]
    warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------


def _s(v) -> str:
    """Cell value → trimmed string; None → ''."""
    if v is None:
        return ""
    return str(v).strip()


def _nonblank(row: tuple) -> bool:
    return any(_s(c) != "" for c in row)


def _find_label_row(ws: Worksheet, label_needle: str, max_row: int = 100, search_cols: int = 10) -> int | None:
    """Find the first row whose any cell (within `search_cols`) strip-equals the needle."""
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row, max_col=search_cols, values_only=True),
        start=1,
    ):
        for cell in row:
            if _s(cell) == label_needle:
                return r_idx
    return None


def _cell_right_of_label(ws: Worksheet, label: str, max_row: int = 100, search_cols: int = 10) -> str:
    """Find the label in any cell, return the non-empty cell to its immediate right.

    Tolerates multiple label hits (takes first). Returns '' if not found.
    """
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=search_cols, values_only=True):
        for c_idx, cell in enumerate(row[:-1]):
            if _s(cell) == label:
                # scan right until we find non-empty
                for follow in row[c_idx + 1:]:
                    if _s(follow) != "":
                        return _s(follow)
                return ""
    return ""


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------


def parse_project_info(ws: Worksheet, spec: ParsedSpec) -> None:
    spec.project.name_zh = _cell_right_of_label(ws, "遊戲名稱(中文)")
    spec.project.name_en = _cell_right_of_label(ws, "遊戲名稱(英文)")
    spec.project.game_id = _cell_right_of_label(ws, "GAME ID")
    spec.project.short_name = _cell_right_of_label(ws, "縮寫")

    for field_name, attr in (
        ("遊戲名稱(中文)", "name_zh"),
        ("GAME ID", "game_id"),
        ("縮寫", "short_name"),
    ):
        if not getattr(spec.project, attr):
            spec.warnings.append(f"project_info: 找不到欄位 {field_name!r}")


def _parse_board_size(board_str: str) -> tuple[int | None, int | None]:
    """Parse '3X5' → (ROW=3, COL=5).  Also accepts '3x5', '3×5', ' 3 X 5 '."""
    if not board_str:
        return None, None
    m = re.search(r"(\d+)\s*[xX×]\s*(\d+)", board_str)
    if not m:
        return None, None
    return int(m.group(1)), int(m.group(2))


def _classify_pay_mode(lines_str: str) -> str:
    """'243 WAYS' / '1024 WAYS' → SCATTER_PAY; '10 LINES' → LINE_PAY."""
    if not lines_str:
        return "UNKNOWN"
    s = lines_str.upper()
    if "WAY" in s:
        return "SCATTER_PAY"
    if "LINE" in s or "線" in lines_str:
        return "LINE_PAY"
    if "CLUSTER" in s:
        return "CLUSTER_PAY"
    return "UNKNOWN"


def parse_basic_header(ws: Worksheet, spec: ParsedSpec) -> None:
    board = _cell_right_of_label(ws, "盤面大小", max_row=50)
    lines = _cell_right_of_label(ws, "中獎線數", max_row=50)
    sym_notes = _cell_right_of_label(ws, "符號種類", max_row=50)

    spec.row, spec.col = _parse_board_size(board)
    spec.pay_mode = _classify_pay_mode(lines)
    spec.symbol_notes = sym_notes
    # Uniform board: generate layout string like "3x3x3x3x3"
    if spec.row and spec.col:
        spec.board_layout = "x".join([str(spec.row)] * spec.col)

    if spec.row is None or spec.col is None:
        spec.warnings.append(
            f"basic: 無法解析盤面大小 (原值 {board!r})。請確認『基本規格』sheet 的『盤面大小』值格式為 'ROW×COL' (如 3X5)"
        )
    if spec.pay_mode == "UNKNOWN":
        spec.warnings.append(
            f"basic: 無法辨識中獎線數 {lines!r} 的賠付模式，fallback=UNKNOWN"
        )


_CATEGORY_FROM_DISPLAY = [
    # lower-priority displays first so specific matches win
    ("SCATTER", "scatter"),
    ("WILD", "wild"),
    ("神秘", "special"),
    ("金幣", "special"),
]


# Whitelist of safe display → enum-name aliases.
# Rules:
#   - Only map displays we are CERTAIN about; preserve Symbol_NN for everything
#     else (e.g. 盆栽, 扇子, 猛虎, 神龍 are game-specific Chinese names that we
#     refuse to auto-translate — wrong English could break downstream).
#   - SCATTER(X) → ScatterX via _scatter_variant_alias() so we don't have to
#     enumerate every variant.
#   - Keys are uppercase-compared so "wild" / "Wild" / "WILD" all match.
_SYMBOL_ALIAS_MAP: dict[str, str] = {
    "9": "Nine",
    "10": "Ten",
    "J": "J",
    "Q": "Q",
    "K": "K",
    "A": "A",
    "WILD": "Wild",
    "SCATTER": "Scatter",
}


def _scatter_variant_alias(display: str) -> str | None:
    """Map 'SCATTER(猛虎)' → 'ScatterTiger' only for a small whitelist of
    well-known Mandarin game symbols we've seen in UK slots. Return None if
    we can't be confident."""
    # only handle forms like SCATTER(...)
    m = re.match(r"SCATTER\(([^)]+)\)", display.strip(), flags=re.IGNORECASE)
    if not m:
        return None
    inner = m.group(1).strip()
    # whitelist — add more as we encounter new UK slot Chinese variants
    inner_map: dict[str, str] = {
        "猛虎": "Tiger",
        "神龍": "Dragon",
        "鯉魚": "Koi",
        "虎": "Tiger",
        "龍": "Dragon",
        "魚": "Koi",
    }
    if inner in inner_map:
        return f"Scatter{inner_map[inner]}"
    # if inner is already ASCII (e.g. SCATTER(GOLD)) → ScatterGold
    if re.fullmatch(r"[A-Za-z0-9_]+", inner):
        return f"Scatter{inner.capitalize()}"
    return None


def _resolve_enum_name(raw: str, display: str) -> str:
    """Decide the TypeScript enum name.

    Preference order:
      1. Scatter variant match (e.g. SCATTER(猛虎) → ScatterTiger)
      2. Direct whitelist hit on display (e.g. '9' → 'Nine', 'WILD' → 'Wild')
      3. Fall back to raw 系統命名 column value (e.g. 'Symbol_06') — this is
         always a valid TypeScript identifier and preserves the original
         ordering.

    We deliberately DO NOT translate game-specific Chinese symbols (盆栽,
    扇子, 猛虎 etc.) to avoid brittle/ambiguous English names.
    """
    d = (display or "").strip()
    if not d:
        return raw or "UnnamedSymbol"
    # 1. scatter variant
    scatter_alias = _scatter_variant_alias(d)
    if scatter_alias:
        return scatter_alias
    # 2. direct whitelist
    alias = _SYMBOL_ALIAS_MAP.get(d.upper())
    if alias:
        return alias
    # 3. fallback
    return raw or f"Symbol_Unknown_{d}"


def _is_server_only(side_col_b: str) -> bool:
    """Detect whether this symbol row falls under the 'Server用' sticky side tag."""
    side = side_col_b or ""
    side_normalized = re.sub(r"\s+", "", side)
    return (
        "SERVER" in side_normalized.upper()
        or "伺服器" in side_normalized
        or side_normalized == "Server用"
    )


def _classify_symbol(display: str, side_col_b: str, idx: int, total_low_hint: int | None) -> str:
    """Best-effort classification.

    Rules:
      1. server-only rows → classified to 'special' by category, and `server_only=True`
         is recorded separately on SymbolRow by the caller. This keeps the set of
         categories aligned with `_input-format.md` (low/high/wild/scatter/special)
         while preserving the 'server-only' semantic elsewhere. (See E2E-1 fix)
      2. col B tag '特殊' + display contains SCATTER/WILD → that
      3. col B tag '特殊' + otherwise → special
      4. col B tag '一般' → low/high by display heuristic
         - display in {9,10,J,Q,K,A} → low
         - everything else → high
    """
    side = side_col_b or ""
    side_normalized = re.sub(r"\s+", "", side)
    disp_upper = display.upper()
    # (1) Server-only rows collapse to 'special' for the skill's 5-category schema.
    if _is_server_only(side):
        return "special"
    if "特殊" in side_normalized:
        for keyword, cat in _CATEGORY_FROM_DISPLAY:
            if keyword in disp_upper:
                return cat
        return "special"
    # 一般 / default
    if disp_upper in {"9", "10", "J", "Q", "K", "A"} or display in {"9", "10"}:
        return "low"
    return "high"


def parse_symbol_table(ws: Worksheet, spec: ParsedSpec) -> None:
    """Walk rows to find '■Symbol賠付表' header, then parse rows until next '■'."""
    start_row = _find_label_row(ws, "■Symbol賠付表  (Win=Total Bet*Pay)", max_row=300)
    if start_row is None:
        # try looser match
        for r_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=300, max_col=10, values_only=True),
            start=1,
        ):
            for cell in row:
                txt = _s(cell)
                if txt.startswith("■Symbol") or txt.startswith("■符號"):
                    start_row = r_idx
                    break
            if start_row:
                break
    if start_row is None:
        spec.warnings.append("symbols: 找不到『■Symbol賠付表』區塊，無法解析符號清單")
        return

    # data rows start 2 rows after the "編號 | 系統命名 | SYMBOL樣式 | ..." header
    current_side = ""  # 一般 / 特殊 / Server用 (sticky between rows)
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=start_row + 1, max_row=start_row + 80, max_col=10, values_only=True),
        start=start_row + 1,
    ):
        cells = [_s(c) for c in row]
        # abort if we hit the next "■" section
        if any(c.startswith("■") for c in cells):
            break

        col_b, col_c, col_d, col_e = cells[1], cells[2], cells[3], cells[4]
        # normalize the side tag: Excel cells sometimes contain an embedded newline
        # like "Server\n用" due to in-cell wrapping; collapse all whitespace so the
        # sticky state machine picks it up correctly.
        col_b_normalized = re.sub(r"\s+", "", col_b)
        if col_b_normalized in ("一般", "特殊", "Server用"):
            current_side = col_b_normalized
        # expect col_c to be numeric symbol index
        if not re.fullmatch(r"\d+", col_c):
            continue
        idx = int(col_c)
        raw_enum = col_d  # Symbol_00 / Symbol_11 etc
        display = col_e
        if not raw_enum:
            continue
        category = _classify_symbol(display, current_side, idx, None)
        server_only = _is_server_only(current_side)
        # Alias lookup: try to promote well-known symbols ('9'→Nine, 'WILD'→Wild,
        # 'SCATTER(猛虎)'→ScatterTiger) while keeping non-whitelisted Chinese
        # names (盆栽/扇子/猛虎/神龍) as Symbol_NN for safety.
        enum_name = _resolve_enum_name(raw_enum, display)
        spec.symbols.append(
            SymbolRow(
                index=idx,
                enum_name=enum_name,
                raw_enum_name=raw_enum,
                display=display,
                category=category,
                server_only=server_only,
            )
        )

    if not spec.symbols:
        spec.warnings.append("symbols: 已找到賠付表 header，但解析不到任何 symbol 資料列")


def parse_cheat_keys(ws: Worksheet, spec: ParsedSpec) -> None:
    start_row = None
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=300, max_col=10, values_only=True),
        start=1,
    ):
        for cell in row:
            txt = _s(cell)
            if txt.startswith("■特色展示碼"):
                start_row = r_idx
                break
        if start_row:
            break
    if start_row is None:
        spec.warnings.append("cheat_keys: 找不到『■特色展示碼』區塊，Demo_Module CHEAT_KEY 將使用預設值")
        return

    # skip the header row ("效果 | 編號 | 備註")
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=start_row + 1, max_row=start_row + 80, max_col=10, values_only=True),
        start=start_row + 1,
    ):
        cells = [_s(c) for c in row]
        if any(c.startswith("■") for c in cells):
            break
        effect = cells[1]
        code = cells[2]
        note = cells[3] if len(cells) > 3 else ""
        # header row has effect="效果"
        if effect in ("", "效果"):
            continue
        if not re.fullmatch(r"-?\d+", code):
            continue
        spec.cheat_keys.append(CheatKeyRow(effect=effect, code=code, note=note))


def parse_audio(ws: Worksheet, spec: ParsedSpec) -> None:
    """Walk 音樂音效規格表. Each section header row has col B = section name
    and col C = 'LOOP'. Data rows have col A = index number."""
    current_section = ""
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=200, max_col=10, values_only=True),
        start=1,
    ):
        cells = [_s(c) for c in row]
        col_a, col_b, col_c, col_g = cells[0], cells[1], cells[2] if len(cells) > 2 else "", cells[6] if len(cells) > 6 else ""
        note_col = cells[5] if len(cells) > 5 else ""

        # Section header detection: col B looks like a name, col C == 'LOOP'
        if col_c == "LOOP" and col_b and col_a == "":
            current_section = col_b
            continue

        # skip pure header rows
        if not col_g:
            continue
        if col_b in ("項目", "", "LOOP"):
            continue
        # 檔名 col_g: allow multi-line (the xlsx sometimes stuffs 3 names separated by \n)
        filenames = [fn.strip() for fn in col_g.replace("\r", "").split("\n") if fn.strip()]
        for fn in filenames:
            spec.audio_clips.append(
                AudioClip(section=current_section or "未分類", item_zh=col_b, filename=fn, note=note_col)
            )


# ---------------------------------------------------------------------------
# Performance-flow hint scanner
# ---------------------------------------------------------------------------


# Map of sheet name → (hint_kind, one-line description).
# We walk this table against the xlsx's sheet list and record any match, so the
# final Game_Spec.md's Performance_Flow section tells the designer what extra
# information the xlsx contains. We deliberately DO NOT try to parse these
# sheets into step lists — empirically they contain art-animation details
# (Spine nodes / durations / effect-layer names) or state-combo tags, not
# sequenced state-machine steps. Trying to auto-derive steps from them would
# produce a brittle, misleading Performance_Flow.
_PERFORMANCE_HINT_SHEETS: list[tuple[str, str, str]] = [
    (
        "三色幣主場狀態設定",
        "state-combo-tag",
        "列出 Scatter 組合類型 tag（如 Scatter_3MG / Scatter_1B / Scatter_2BR），"
        "非線性流程步驟。若 codegen skill 需區分多 Scatter 組合，這些 tag 可當 proto enum。",
    ),
    (
        "Scatter_JP動畫狀態",
        "state-combo-tag",
        "Scatter JP 動畫狀態 tag；空 sheet 或極少內容時代表此特色未啟用。",
    ),
    (
        "前導動畫串接",
        "art-animation",
        "前導動畫（GameIntro / Reel 出場）的 Spine 節點 + 動畫時長描述。"
        "屬美術交付細節，不是狀態機步驟。",
    ),
    (
        "NearWin動畫",
        "art-animation",
        "NearWin 特效層名（如 FX_NearWin / FX_NearWin_Back）；供美術端放節點參考。",
    ),
    (
        "FG_Declare宣告",
        "art-animation",
        "FG 宣告畫面的 Spine / UI 細節，若空代表此流程走共用模板。",
    ),
    (
        "NodeTree",
        "scene-structure",
        "整棵 Scene 節點樹。對應 Cocos Scene 建構（不進 Game_Spec，但 codegen 階段二可用於輔助 Scene 建置）。",
    ),
    (
        "Prefab資源",
        "art-manifest",
        "Prefab 清單；屬 ART_ASSET_MANIFEST 範疇，不影響 Performance_Flow。",
    ),
]


def scan_performance_hints(wb: openpyxl.workbook.workbook.Workbook, spec: ParsedSpec) -> None:
    """Don't parse these sheets — only record their presence + size, so the
    resulting Game_Spec.md can tell downstream agents / designers which extra
    context is in the xlsx.
    """
    for sheet_name, kind, summary in _PERFORMANCE_HINT_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # rough size: count rows that have any non-empty cell
        nonempty_rows = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 600), max_col=min(ws.max_column, 30), values_only=True):
            if any(_s(c) != "" for c in row):
                nonempty_rows += 1
        spec.performance_hints.append(
            {
                "sheet": sheet_name,
                "kind": kind,
                "summary": summary,
                "row_count": nonempty_rows,
            }
        )


def derive_feature_flags(spec: ParsedSpec) -> None:
    # HAS_FREE_GAME: any symbol is scatter OR audio section contains 'FG'
    has_scatter_symbol = any(s.category == "scatter" for s in spec.symbols)
    has_fg_audio = any(
        "FG" in c.section or c.item_zh.startswith("FG") or c.filename.lower().startswith("fg_")
        for c in spec.audio_clips
    )
    if has_scatter_symbol or has_fg_audio:
        spec.has_free_game = True

    # HAS_JACKPOT: audio contains Grand, cheat has GRAND/JP keyword,
    # OR symbol table has JP_ prefixed symbols (symbol-based JP like COLLECT 收分)
    has_jp_audio = any(
        "Grand" in c.item_zh or "Grand" in c.filename or "Grand" in c.note
        for c in spec.audio_clips
    )
    has_jp_cheat = any(
        "JP" in ck.effect.upper() or "GRAND" in ck.effect.upper() for ck in spec.cheat_keys
    )
    has_jp_symbol = any(
        s.display.upper().startswith("JP_") or s.enum_name.upper().startswith("JP")
        or "JACKPOT" in s.display.upper()
        for s in spec.symbols
    )
    if has_jp_audio or has_jp_cheat or has_jp_symbol:
        spec.has_jackpot = True

    # HAS_BUY_BONUS: cheat-key or symbol table hint
    if any("BUY" in ck.effect.upper() or "購買" in ck.effect for ck in spec.cheat_keys):
        spec.has_buy_bonus = True

    # CUSTOM_FEATURES: detect non-standard mechanics from audio/cheat/symbol/performance hints
    _standard_keywords = {"FG", "FREE", "SCATTER", "WILD", "GRAND", "MAJOR", "MINOR", "MINI",
                          "JP", "JACKPOT", "BUY", "BONUS", "BGM", "SPIN", "WIN", "STOP",
                          "NEAR", "BASE", "NORMAL", "BUTTON", "REEL"}
    custom_seen: set[str] = set()
    # From audio sections that hint at custom mechanics
    for c in spec.audio_clips:
        section_upper = c.section.upper() if c.section else ""
        if section_upper and not any(kw in section_upper for kw in _standard_keywords):
            label = c.section.strip()
            if label and label not in custom_seen:
                custom_seen.add(label)
    # From cheat keys with non-standard effects
    for ck in spec.cheat_keys:
        effect_upper = ck.effect.upper()
        if not any(kw in effect_upper for kw in _standard_keywords):
            label = ck.effect.strip()
            if label and label not in custom_seen:
                custom_seen.add(label)
    # From performance_hints with unique kind
    for ph in spec.performance_hints:
        kind = ph.get("kind", "")
        if kind and kind.upper() not in {"STANDARD", "FREE_GAME", "BASE_GAME", "RESPIN"}:
            if kind not in custom_seen:
                custom_seen.add(kind)
    spec.custom_features = sorted(custom_seen)


# ---------------------------------------------------------------------------
# Markdown emitter
# ---------------------------------------------------------------------------


def _audio_key_from_item(item_zh: str, filename: str, section: str) -> str:
    """Build an AudioClips key that reads nicely in TS code. Prefer filename basename
    over zh translation because filename is already an identifier."""
    if filename:
        return filename  # spec-adapter leaves it in raw form; codegen skill will PascalCase if needed
    # fallback: strip spaces from zh
    return re.sub(r"\s+", "_", item_zh)


def emit_markdown(spec: ParsedSpec) -> str:
    lines: list[str] = []
    P = spec.project

    title = (P.name_zh or P.name_en or "Unknown Slot").strip()
    lines.append(f"# Game_Spec — {title}")
    lines.append("")
    lines.append("> 本檔由 `uk-slot-spec-adapter` 從 xlsx 自動產生。")
    lines.append("> 作為 `slot-game-codegen-skill` 階段一（Game_Summary_File 生成）的輸入。")
    lines.append("> 若有 ⚠ 警示，代表對應欄位在原 xlsx 中無法被自動解析，請人工補齊。")
    lines.append("")

    # ----- Parse warnings first (hard to miss) -----
    if spec.warnings:
        lines.append("## ⚠ Parse Warnings")
        lines.append("")
        for w in spec.warnings:
            lines.append(f"- {w}")
        lines.append("")

    # ----- Project info -----
    lines.append("## 專案資訊")
    lines.append("")
    lines.append(f"- 遊戲名稱（中文）：{P.name_zh or '⚠ 未找到'}")
    lines.append(f"- 遊戲名稱（英文）：{P.name_en or '⚠ 未找到'}")
    lines.append(f"- GAME ID：{P.game_id or '⚠ 未找到'}")
    lines.append(f"- 縮寫（short name）：{P.short_name or '⚠ 未找到'}")
    lines.append("")

    # ----- Reel config -----
    lines.append("## 轉輪配置")
    lines.append("")
    lines.append(f"- COL: {spec.col if spec.col is not None else '⚠ 未解析'}")
    lines.append(f"- ROW: {spec.row if spec.row is not None else '⚠ 未解析'}")
    if spec.col and spec.row:
        if spec.board_layout and "x" in spec.board_layout:
            clean_bl = re.split(r"[、，,\(（]", spec.board_layout)[0].strip()
            parts = re.split(r"[xX×]", clean_bl)
            if len(parts) > 2 and all(p.strip().isdigit() for p in parts):
                # Variable board: sum of per-column rows
                total = sum(int(p.strip()) for p in parts)
            else:
                # Simple ROWxCOL format: uniform board
                total = spec.col * spec.row
        else:
            total = spec.col * spec.row
        lines.append(f"- FULL_PLATE_NUM: {total}（各列 row 加總）")
    if spec.board_layout:
        lines.append(f"- BoardLayout: `{spec.board_layout}`")
    lines.append(f"- PayMode: `{spec.pay_mode}`")
    if spec.symbol_notes:
        lines.append(f"- 符號種類（原文）：{spec.symbol_notes}")
    lines.append("")

    # ----- Symbols -----
    lines.append("## 符號定義")
    lines.append("")
    if not spec.symbols:
        lines.append("⚠ 未解析到任何符號，請檢查原 xlsx 的『■Symbol賠付表』區塊。")
    else:
        lines.append("| 列舉名稱 | 原命名 | 類型 | 說明 | 備註 |")
        lines.append("|---|---|---|---|---|")
        for s in spec.symbols:
            note_bits = [f"idx={s.index}"]
            if s.server_only:
                note_bits.append("Server only")
            if s.enum_name != s.raw_enum_name:
                note_bits.append(f"alias ← {s.display!r}")
            note_str = "; ".join(note_bits)
            # If alias was used, 原命名 column shows Symbol_NN as a cross-reference;
            # otherwise it's the same as enum_name.
            raw_display = s.raw_enum_name if s.raw_enum_name != s.enum_name else "—"
            lines.append(f"| {s.enum_name} | {raw_display} | {s.category} | {s.display} | {note_str} |")
    lines.append("")

    # ----- Features -----
    lines.append("## 特色玩法")
    lines.append("")
    lines.append(f"- HAS_FREE_GAME: **{spec.has_free_game}**")
    lines.append(f"- HAS_JACKPOT: **{spec.has_jackpot}**")
    lines.append(f"- HAS_BUY_BONUS: **{spec.has_buy_bonus}**")
    lines.append("")
    if spec.custom_features:
        lines.append("## Custom Features")
        lines.append("")
        lines.append("> 以下為 adapter 偵測到的非標準特色玩法（不在 FG/JP/BuyBonus 範疇），")
        lines.append("> 需 Step 3.10.8 人工確認是否實作。")
        lines.append("")
        for cf in spec.custom_features:
            lines.append(f"- {cf}")
        lines.append("")
    lines.append("> 以上旗標為 adapter 依符號表 / 音效表 / CHEAT_KEY 關鍵字推導，")
    lines.append("> 若與設計預期不同請於 Game_Summary_File 階段手動覆寫。")
    lines.append("")

    # ----- Audio -----
    lines.append("## 音效需求清單")
    lines.append("")
    if not spec.audio_clips:
        lines.append("⚠ 未解析到任何音效項目，請檢查原 xlsx 的『音樂音效規格表』sheet。")
    else:
        lines.append("| 鍵名(Key / FileName) | 檔名 | 分區 | 說明 |")
        lines.append("|---|---|---|---|")
        for c in spec.audio_clips:
            key = _audio_key_from_item(c.item_zh, c.filename, c.section)
            note = (c.note or "").replace("\n", " ").replace("|", "／")
            item = (c.item_zh or "").replace("|", "／")
            lines.append(f"| {key} | {c.filename} | {c.section} | {item}｜{note} |")
    lines.append("")

    # ----- CHEAT_KEY -----
    lines.append("## CHEAT_KEY（特色展示碼）")
    lines.append("")
    if not spec.cheat_keys:
        lines.append("⚠ 未解析到 CHEAT_KEY 清單，或原 xlsx 未提供。")
    else:
        lines.append("| 效果 | 編號 | 備註 |")
        lines.append("|---|---|---|")
        for ck in spec.cheat_keys:
            note = (ck.note or "").replace("\n", " ").replace("|", "／")
            lines.append(f"| {ck.effect} | {ck.code} | {note} |")
    lines.append("")

    # ----- Performance flow (presence-only reporting, intentionally not parsed) -----
    lines.append("## 演出流程（Performance_Flow）")
    lines.append("")
    lines.append("> **adapter 設計決定**：不自動推導 Performance_Flow 步驟清單。")
    lines.append(">")
    lines.append("> 原因：xlsx 內「三色幣主場狀態設定 / 前導動畫串接 / NearWin動畫 / FG_Declare宣告」")
    lines.append("> 等 sheet 實測為**美術動畫細節或狀態組合 tag**，不是線性狀態機步驟；")
    lines.append("> 硬解會產生誤導性的 Performance_Flow，不如不給。")
    lines.append(">")
    lines.append("> **預設行為**：codegen skill 應採用 `Standard_Flow`（`EFFECT_START → SCATTER_SHOW →")
    lines.append("> AWARD → ROUND_SHOW_END → CHECK_STATE → ROUND_END`）；複合玩法的差異由 CHECK_STATE")
    lines.append("> 內部分支與特色旗標（HAS_FREE_GAME / HAS_JACKPOT / ...）驅動，無須自訂 Performance_Flow。")
    lines.append(">")
    lines.append("> **何時應手動覆寫**：只有 Cascade / Tumble 這類「步驟本身與 Standard 不同」的玩法，")
    lines.append("> 才需要在 Game_Spec 的 `【演出流程】` 章節自訂 step table（格式見 codegen skill 的")
    lines.append("> `_performance-flow-examples.md`）。三幣瑞龍此類複合 FG 玩法**不需要**。")
    lines.append("")

    # Presence map — tell the designer what extra sheets exist in the xlsx
    if spec.performance_hints:
        lines.append("### 原 xlsx 中的流程相關 sheet（adapter 不解析，供設計者參考）")
        lines.append("")
        lines.append("| Sheet 名 | 類型 | 非空行數 | 說明 |")
        lines.append("|---|---|---|---|")
        for h in spec.performance_hints:
            kind = h["kind"]
            summary = h["summary"].replace("\n", " ")
            lines.append(f"| `{h['sheet']}` | {kind} | {h['row_count']} | {summary} |")
        lines.append("")
        lines.append("> 類型說明：")
        lines.append(">")
        lines.append("> - `state-combo-tag`：列特殊狀態組合的命名（如 Scatter 三色幣 7 種組合）。可能對應 proto enum，但不是步驟順序。")
        lines.append("> - `art-animation`：美術動畫的 Spine 節點名 / 時長 / 特效層。屬 `ART_ASSET_MANIFEST` 範疇。")
        lines.append("> - `scene-structure`：Cocos Scene 節點樹；可用於 codegen 階段二的 Scene 建置輔助。")
        lines.append("> - `art-manifest`：Prefab 清單；屬 `ART_ASSET_MANIFEST` 範疇。")
        lines.append("")
    else:
        lines.append("> 原 xlsx 中未偵測到任何演出相關 sheet。採用 Standard_Flow。")
        lines.append("")

    # ----- Footer -----
    lines.append("---")
    lines.append("")
    lines.append("### Adapter metadata")
    lines.append("")
    lines.append(f"- symbol count: {len(spec.symbols)}")
    lines.append(f"- cheat-key count: {len(spec.cheat_keys)}")
    lines.append(f"- audio-clip count: {len(spec.audio_clips)}")
    lines.append(f"- warnings: {len(spec.warnings)}")
    lines.append("")

    return "\n".join(lines) + "\n"


# ---------------------------------------------------------------------------
# Layout B parsers (Eye Strike / Gold Blitz style: numbered sheets like
# "2.遊戲概述", "6.圖騰設計", "9.音樂音效(暫)")
# ---------------------------------------------------------------------------


def _parse_board_size_variable(board_str: str) -> tuple[int | None, int | None, str]:
    """Parse variable-height boards like '5x4x4x4x4x5' → (col=6, max_row=5, raw).
    Also handles simple '3X5' → (col=5, row=3, raw).
    Strips notes after '、', '(', '(' before parsing."""
    if not board_str:
        return None, None, ""
    raw = board_str.strip()
    # Strip trailing notes (e.g. "3x5、5x5 (其中...)" → "3x5")
    clean = re.split(r"[、，,\(（]", raw)[0].strip()
    # Try variable format first: 5x4x4x4x4x5 (all parts must be digits)
    parts = re.split(r"[xX×]", clean)
    if len(parts) > 2 and all(p.strip().isdigit() for p in parts):
        col = len(parts)
        max_row = max(int(p.strip()) for p in parts)
        return col, max_row, raw
    # Simple format
    m = re.search(r"(\d+)\s*[xX×]\s*(\d+)", clean)
    if m:
        return int(m.group(2)), int(m.group(1)), raw  # COL, ROW
    return None, None, raw


def parse_layout_b_overview(ws: Worksheet, spec: ParsedSpec) -> None:
    """Parse '2.遊戲概述' sheet for board size, pay mode, and ODDS table symbols."""
    # Board size: find row with '盤面'
    for row in ws.iter_rows(min_row=1, max_row=30, max_col=10, values_only=True):
        cells = [_s(c) for c in row]
        if "盤面" in cells[1]:
            board_val = cells[3] or cells[2]
            col, max_row, raw = _parse_board_size_variable(board_val)
            spec.col = col
            spec.row = max_row
            spec.board_layout = raw
            if col is None:
                spec.warnings.append(f"basic: 無法解析盤面大小 (原值 {board_val!r})")
            break

    # Pay mode: find row with '線數' or '對獎方式' or '對線'
    for row in ws.iter_rows(min_row=1, max_row=30, max_col=10, values_only=True):
        cells = [_s(c) for c in row]
        if "線數" in cells[1] or "對獎" in cells[1] or "對線" in cells[1]:
            spec.pay_mode = _classify_pay_mode(cells[3] or cells[2])
            break

    # Game name from row with '基本規格說明'
    for row in ws.iter_rows(min_row=1, max_row=15, max_col=10, values_only=True):
        cells = [_s(c) for c in row]
        if "基本規格說明" in cells[1]:
            spec.project.name_zh = ""  # will be filled from audio sheet
            break

    # Feature flags from key-value rows
    for row in ws.iter_rows(min_row=1, max_row=20, max_col=10, values_only=True):
        cells = [_s(c) for c in row]
        label = cells[1].lower() if len(cells) > 1 else ""
        val = (cells[2] or "").lower() if len(cells) > 2 else ""
        if "free game" in label and ("true" in val or "有" in val):
            spec.has_free_game = True
        if "jp" == label.strip() and ("true" in val or "有" in val):
            spec.has_jackpot = True
        if "buy bonus" in label and ("true" in val or "有" in val):
            spec.has_buy_bonus = True


def parse_layout_b_symbols(ws: Worksheet, spec: ParsedSpec) -> None:
    """Parse symbol sheet. Supports multiple formats:
    - Format 1 (Eye Strike): col C = numeric ID, col D = name
    - Format 2 (Wrath of Thunder): col C = symbol name (M1/A/K), col D = size "1x1"
    - Format 3 (FAR WEST): col B = symbol name, col D = numeric ID
    - Format 4 (Leprechaun's Pots): col D = name, col F = symbol code (M1/H1), col G = area "1x1"
    """
    current_section = "一般"  # 一般符號 / 特殊符號
    auto_idx = 0

    # Detect Format 4: header row has '名稱' and '符號' and '面積'
    is_format4 = False
    format4_cols = {}  # name_col, sym_col, area_col
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, max_col=12, values_only=True), 1):
        cells = [_s(c) for c in row]
        if "名稱" in cells and "符號" in cells and "面積" in cells:
            is_format4 = True
            format4_cols = {
                "name": cells.index("名稱"),
                "sym": cells.index("符號"),
                "area": cells.index("面積"),
            }
            # Also grab SymID col if present
            if "SymID" in cells:
                format4_cols["symid"] = cells.index("SymID")
            break

    if is_format4:
        nc = format4_cols["name"]
        sc = format4_cols["sym"]
        sid_col = format4_cols.get("symid")
        for row in ws.iter_rows(min_row=1, max_row=120, max_col=12, values_only=True):
            cells = [_s(c) for c in row] + [""] * 12
            if "特殊符號" in cells[1] or "特殊" in cells[1]:
                current_section = "特殊"
                continue
            if "一般符號" in cells[1] or "一般" in cells[1]:
                current_section = "一般"
                continue
            sym_code = cells[sc]
            display_name = cells[nc]
            if not sym_code or not re.fullmatch(r"[A-Za-z]\w*", sym_code):
                continue
            if sym_code in ("符號", "名稱", "面積", "說明", "None"):
                continue
            if sid_col and cells[sid_col] and re.fullmatch(r"\d+", cells[sid_col]):
                idx = int(cells[sid_col])
            else:
                idx = auto_idx
            auto_idx += 1
            disp_upper = sym_code.upper()
            if "WILD" in disp_upper or "WW" in disp_upper:
                category = "wild"
            elif "SCATTER" in disp_upper or display_name and "scatter" in display_name.lower():
                category = "scatter"
            elif current_section == "特殊":
                category = "special"
            elif disp_upper in {"9", "10", "J", "Q", "K", "A"}:
                category = "low"
            else:
                category = "high"
            enum_name = _resolve_enum_name(f"Symbol_{idx:02d}", sym_code)
            spec.symbols.append(SymbolRow(
                index=idx, enum_name=enum_name, raw_enum_name=f"Symbol_{idx:02d}",
                display=f"{sym_code} ({display_name})" if display_name else sym_code,
                category=category, server_only=False,
            ))
        return

    for row in ws.iter_rows(min_row=1, max_row=120, max_col=10, values_only=True):
        cells = [_s(c) for c in row]
        # Section markers
        if "特殊符號" in cells[1] or "特殊" in cells[1]:
            current_section = "特殊"
            continue
        if "一般符號" in cells[1] or "一般" in cells[1]:
            current_section = "一般"
            continue

        # Format 1: col C is numeric ID (but NOT if col D looks like area "1x1")
        col_c = cells[2]
        col_d = cells[3] if len(cells) > 3 else ""
        col_b = cells[1]
        is_area_format = bool(re.fullmatch(r"\d+x\d+", col_d))
        if re.fullmatch(r"\d+", col_c) and not is_area_format:
            idx = int(col_c)
            display = col_d.split("\n")[0].strip() if col_d else ""
        # Format 2: col C is symbol name, col D is area "1x1"
        elif re.fullmatch(r"[A-Za-z]\w*", col_c) and col_c not in ("參考", "符號", "面積", "外框", "色調", "說明", "None", "True", "False") and is_area_format:
            display = col_c
            idx = auto_idx
            auto_idx += 1
        # Format 3: col B is symbol name, col D is numeric ID (FAR WEST style)
        elif re.fullmatch(r"[A-Za-z]\w*", col_b) and col_b not in ("符號", "名稱", "參考", "一般符號", "特殊符號"):
            display = col_b
            if col_d and re.fullmatch(r"\d+", col_d):
                idx = int(col_d)
            else:
                idx = auto_idx
            auto_idx += 1
        else:
            continue

        # Classify
        disp_upper = display.upper()
        if "WILD" in disp_upper:
            category = "wild"
        elif "SCATTER" in disp_upper:
            category = "scatter"
        elif current_section == "特殊":
            category = "special"
        elif disp_upper in {"9", "10", "J", "Q", "K", "A"}:
            category = "low"
        else:
            category = "high"

        enum_name = _resolve_enum_name(f"Symbol_{idx:02d}", display)
        spec.symbols.append(SymbolRow(
            index=idx,
            enum_name=enum_name,
            raw_enum_name=f"Symbol_{idx:02d}",
            display=display,
            category=category,
            server_only=False,
        ))



def _parse_inline_symbols(ws: Worksheet, spec: ParsedSpec) -> None:
    """Parse symbols from inline Odds Table in overview sheet (FAR WEST style).
    Format: | 名稱 | 符號 | SymID | 2連線 | 3連線 | ... |"""
    in_odds = False
    current_section = "一般"
    for row in ws.iter_rows(min_row=1, max_row=80, max_col=10, values_only=True):
        cells = [_s(c) for c in row]
        # Detect Odds Table start
        if "Odds" in cells[0]:
            in_odds = True
            continue
        if not in_odds:
            continue
        # Section markers
        if "特殊符號" in cells[1] or "特殊" in cells[1]:
            current_section = "特殊"
            continue
        if "一般符號" in cells[1]:
            current_section = "一般"
            continue
        # Symbol row: col B=名稱, col C=符號代碼, col D=SymID
        name = cells[1]
        sym_code = cells[2]
        sym_id = cells[3]
        if not name or not sym_code or name in ("名稱", "符號", "一般符號", "特殊符號"):
            continue
        idx = int(sym_id) if sym_id and re.fullmatch(r"\d+", sym_id) else len(spec.symbols)
        disp_upper = sym_code.upper()
        if "WILD" in disp_upper or "WW" in disp_upper:
            category = "wild"
        elif "SCATTER" in disp_upper or "BS" in disp_upper or "FS" in disp_upper:
            category = "scatter"
        elif current_section == "特殊":
            category = "special"
        elif disp_upper in {"9", "10", "J", "Q", "K", "A"}:
            category = "low"
        else:
            category = "high"
        enum_name = _resolve_enum_name(f"Symbol_{idx:02d}", sym_code)
        spec.symbols.append(SymbolRow(
            index=idx, enum_name=enum_name, raw_enum_name=f"Symbol_{idx:02d}",
            display=f"{sym_code} ({name})", category=category, server_only=False,
        ))


def parse_layout_b_audio(ws: Worksheet, spec: ParsedSpec) -> None:
    """Parse '9.音樂音效(暫)' sheet. Header row has '編號' in col A."""
    # Find game name
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=10, values_only=True):
        cells = [_s(c) for c in row]
        if "機種名稱" in cells[1]:
            spec.project.name_en = cells[3] or cells[2]
            if not spec.project.name_zh:
                spec.project.name_zh = spec.project.name_en
            break

    # Find header row — try exact match first, then fuzzy fallback
    _HEADER_CANDIDATES = {"編號", "流水號", "序號", "項次", "No", "No.", "#"}
    header_row = None
    filename_col = 3  # default: col D
    item_col = 1      # default: col B
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, max_col=15, values_only=True), 1):
        cells = [_s(c) for c in row]
        # Exact match on col A
        if cells[0] in _HEADER_CANDIDATES:
            header_row = r_idx
        # Fuzzy: any cell in this row contains '檔名' AND another contains '品項'/'場景'/'項目'
        elif not header_row:
            has_filename_col = any("檔名" in c for c in cells)
            has_item_col = any(c in ("品項", "運用場景", "項目") or "場景" in c for c in cells)
            if has_filename_col and has_item_col:
                header_row = r_idx
        if header_row and header_row == r_idx:
            # Find '檔名' column
            for ci, c in enumerate(cells):
                if "檔名" in c:
                    filename_col = ci
                    break
            # Find '品項' or '運用場景' or '項目' column
            for ci, c in enumerate(cells):
                if c in ("品項", "運用場景", "項目") or "場景" in c:
                    item_col = ci
                    break
            break
    if header_row is None:
        spec.warnings.append("audio: 找不到音效表 header（嘗試了 '編號/流水號/序號/項次/No' 及模糊比對 '檔名+品項'）")
        return

    current_section = ""
    for row in ws.iter_rows(min_row=header_row + 1, max_row=200, max_col=15, values_only=True):
        cells = [_s(c) for c in row[:15]]
        col_a = cells[0]  # 編號/流水號
        col_item = cells[item_col] if item_col < len(cells) else ""
        col_fn = cells[filename_col] if filename_col < len(cells) else ""

        # Section header: col A empty, col_item has section name
        if not col_a and col_item and not col_fn:
            current_section = col_item
            continue

        # Data row: col A is numeric
        if not re.fullmatch(r"\d+", col_a):
            continue
        if not col_fn:
            continue

        spec.audio_clips.append(AudioClip(
            section=current_section or "未分類",
            item_zh=col_item,
            filename=col_fn,
            note="",
        ))


def _detect_layout_b(wb) -> bool:
    """Detect numbered-sheet layout: Eye Strike / Gold Blitz / Wrath of Thunder / FAR WEST style.
    Has sheets starting with '2.' or '5.' or '6.' or '6-1.' (with or without space after dot)."""
    for name in wb.sheetnames:
        if re.match(r"^[256]\.\s*", name) or re.match(r"^6-1\.", name):
            return True
    return False


def _detect_layout_c(wb) -> bool:
    """Detect 海盜女王 style: has '基礎規則' but NOT '基本規格'."""
    return "基礎規則" in wb.sheetnames and "基本規格" not in wb.sheetnames


def parse_layout_c(ws: Worksheet, spec: ParsedSpec) -> None:
    """Parse 海盜女王 style '基礎規則' sheet.
    Symbol table starts at row with '一般'+'編號'+'SYMBOL樣式'.
    Board info embedded in text like '盤面5 X 6'."""
    # Find board size from text
    for row in ws.iter_rows(min_row=1, max_row=50, max_col=15, values_only=True):
        for cell in row:
            txt = _s(cell)
            m = re.search(r"盤面\s*(\d+)\s*[xX×]\s*(\d+)", txt)
            if m:
                spec.row = int(m.group(1))  # 海盜女王: "盤面5 X 6" → ROW=5? 不對，是 COL=5 ROW=6
                spec.col = int(m.group(1))
                spec.row = int(m.group(2))
                spec.board_layout = "x".join([str(spec.row)] * spec.col)
                break
            # Also try "GAME ID" extraction
            m2 = re.search(r"GAME\s*ID\s*[=＝:：]\s*(\d+)", txt)
            if m2:
                spec.project.game_id = m2.group(1)
        if spec.col:
            break

    # Pay mode from text
    for row in ws.iter_rows(min_row=1, max_row=50, max_col=15, values_only=True):
        for cell in row:
            txt = _s(cell)
            if "多消" in txt or "CLUSTER" in txt.upper():
                spec.pay_mode = "CLUSTER_PAY"
                break
            if "WAY" in txt.upper():
                spec.pay_mode = "SCATTER_PAY"
                break
        if spec.pay_mode:
            break

    # Game name from row 1
    for row in ws.iter_rows(min_row=1, max_row=1, max_col=5, values_only=True):
        cells = [_s(c) for c in row]
        if cells[0]:
            # "海盜女王(難度：5星↑↑↑) " → "海盜女王"
            name = re.sub(r"\(.*\).*", "", cells[0]).strip()
            spec.project.name_zh = name
            break

    # Symbol table: find header row with '編號' + 'SYMBOL樣式'
    header_row = None
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, max_col=10, values_only=True), 1):
        cells = [_s(c) for c in row]
        if "編號" in cells and "SYMBOL樣式" in cells:
            header_row = r_idx
            break
    if not header_row:
        spec.warnings.append("symbols: 找不到 Symbol 表 header")
        return

    current_side = "一般"
    for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + 30, max_col=10, values_only=True):
        cells = [_s(c) for c in row]
        # Side tag in col A
        if cells[0] in ("一般", "特殊"):
            current_side = cells[0]
        # Data: col B = index, col C = display, col G = raw name
        col_b = cells[1]
        col_c = cells[2]
        col_g = cells[6] if len(cells) > 6 else ""
        if not re.fullmatch(r"\d+", col_b):
            continue
        idx = int(col_b)
        display = col_c
        raw_enum = col_g or f"Symbol_{idx:02d}"

        if current_side == "特殊":
            disp_upper = display.upper()
            if "SCATTER" in disp_upper or "海盜旗" in display:
                category = "scatter"
            elif "WILD" in disp_upper:
                category = "wild"
            else:
                category = "special"
        elif display.upper() in {"9", "10", "J", "Q", "K", "A"}:
            category = "low"
        else:
            # Use 層級 col D if available
            level = cells[3] if len(cells) > 3 else ""
            if "低" in level:
                category = "low"
            elif "高" in level:
                category = "high"
            elif "中" in level:
                category = "high"
            else:
                category = "high"

        enum_name = _resolve_enum_name(raw_enum, display)
        spec.symbols.append(SymbolRow(
            index=idx, enum_name=enum_name, raw_enum_name=raw_enum,
            display=display, category=category, server_only=False,
        ))


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def run(xlsx_path: Path, out_path: Path) -> int:
    if not xlsx_path.exists():
        print(f"!! input file not found: {xlsx_path}", file=sys.stderr)
        return 3
    suffix = xlsx_path.suffix.lower()
    if suffix != ".xlsx":
        if suffix in (".docx", ".pdf"):
            # v2 stub: provide actionable guidance rather than a bare error.
            print(
                f"!! {suffix} input is not yet supported by spec_adapter v1.\n"
                f"   Options:\n"
                f"   1. Open the {suffix} in Excel / LibreOffice and Save As .xlsx, then re-run this adapter.\n"
                f"   2. Manually write a Game_Spec.md following the format in\n"
                f"      slot-game-codegen-skill/_game-spec-example.md\n"
                f"   3. Wait for adapter v2 which will add {suffix} parsing "
                f"(python-docx / pypdf — both deps are installed but the parser is not written).",
                file=sys.stderr,
            )
            return 3
        print(f"!! input must be .xlsx (got {xlsx_path.suffix!r})", file=sys.stderr)
        return 3

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    spec = ParsedSpec()

    if _detect_layout_b(wb):
        # Layout B: numbered-sheet style (Eye Strike / Gold Blitz / Wrath of Thunder)
        # Overview sheet: try multiple names
        overview_sheet = None
        for name in wb.sheetnames:
            if re.match(r"^2\.\s*", name):
                overview_sheet = name
                break
        if overview_sheet:
            parse_layout_b_overview(wb[overview_sheet], spec)
        else:
            spec.warnings.append("缺少『2.遊戲概述』或『2.基本規格』sheet")

        # Symbol sheet: try multiple names
        symbol_sheet = None
        for name in wb.sheetnames:
            if re.match(r"^[56]", name) and ("圖騰" in name or "符號" in name or "Symbol" in name.lower()):
                symbol_sheet = name
                break
        if symbol_sheet:
            parse_layout_b_symbols(wb[symbol_sheet], spec)
        else:
            # FAR WEST style: symbols are in the same overview sheet (Odds Table section)
            if overview_sheet:
                _parse_inline_symbols(wb[overview_sheet], spec)
            if not spec.symbols:
                spec.warnings.append("缺少圖騰設計 sheet")

        # Audio: try multiple possible names
        audio_sheet = None
        for name in wb.sheetnames:
            if "音樂音效" in name or "音效" in name:
                audio_sheet = name
                break
        if audio_sheet:
            parse_layout_b_audio(wb[audio_sheet], spec)
        else:
            spec.warnings.append("缺少音樂音效 sheet")

        # Feature flags from sheet names
        for name in wb.sheetnames:
            if "FreeGame" in name or "Free" in name.replace(" ", "") or "_FG" in name:
                spec.has_free_game = True
            if "Buy" in name and "Bonus" in name:
                spec.has_buy_bonus = True
            if "Gold" in name or "Blitz" in name:
                spec.has_jackpot = True
            if "BUY BONUS" in name.upper():
                spec.has_buy_bonus = True

    elif _detect_layout_c(wb):
        # Layout C: 海盜女王 style (基礎規則 with inline symbol table)
        parse_layout_c(wb["基礎規則"], spec)

        # Audio
        audio_sheet = None
        for name in wb.sheetnames:
            if "音樂音效" in name:
                audio_sheet = name
                break
        if audio_sheet:
            parse_layout_b_audio(wb[audio_sheet], spec)
        else:
            spec.warnings.append("缺少音樂音效 sheet")

    else:
        # Layout A: 三幣瑞龍 / 海盜女王 style (13-sheet)
        if "專案資訊" in wb.sheetnames:
            parse_project_info(wb["專案資訊"], spec)
        else:
            spec.warnings.append("缺少『專案資訊』sheet")

        if "基本規格" in wb.sheetnames:
            ws = wb["基本規格"]
            parse_basic_header(ws, spec)
            parse_symbol_table(ws, spec)
            parse_cheat_keys(ws, spec)
        elif "基礎規則" in wb.sheetnames:
            ws = wb["基礎規則"]
            parse_basic_header(ws, spec)
            parse_symbol_table(ws, spec)
            parse_cheat_keys(ws, spec)
        else:
            spec.warnings.append("缺少『基本規格』sheet — 無法產出核心欄位，請檢查 xlsx 結構")

        if "音樂音效規格表" in wb.sheetnames:
            parse_audio(wb["音樂音效規格表"], spec)
        elif "音樂音效規格" in wb.sheetnames:
            parse_audio(wb["音樂音效規格"], spec)
        else:
            spec.warnings.append("缺少『音樂音效規格表』sheet")

    # Performance-flow hint scan (does NOT parse; only records presence + size)
    scan_performance_hints(wb, spec)

    # Sort symbols by their SymID (index) so enum order matches ODDS table.
    # xlsx 內符號列可能不按 SymID 排列（如 WILD 排在中段），不排序會導致
    # enum idx 與 server proto 不對齊。
    if spec.symbols:
        spec.symbols.sort(key=lambda s: s.index)

    # Fallback: derive game name from filename if not parsed
    if not spec.project.name_zh and not spec.project.name_en:
        stem = xlsx_path.stem  # e.g. "3 Leprechaun's Pots_規格書"
        # Remove common suffixes
        name = re.sub(r"[_\s]*(規格書|規格|spec|Spec).*$", "", stem).strip()
        # Remove leading number prefix
        name = re.sub(r"^\d+\s*", "", name).strip()
        if name:
            spec.project.name_en = name

    derive_feature_flags(spec)

    md = emit_markdown(spec)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(md, encoding="utf-8")

    print(f"OK: wrote {out_path}  ({len(md):,} chars, {len(spec.warnings)} warnings)")
    for w in spec.warnings:
        print(f"  ⚠ {w}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="UK slot xlsx → Game_Spec.md adapter")
    ap.add_argument("input", type=Path, help="path to the slot spec xlsx")
    ap.add_argument("output", type=Path, help="destination Game_Spec.md path")
    args = ap.parse_args()
    return run(args.input, args.output)


if __name__ == "__main__":
    sys.exit(main())
